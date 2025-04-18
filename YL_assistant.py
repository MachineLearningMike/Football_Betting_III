
import os
import numpy as np
import pandas as pd
import datetime
import pickle
import json
import re
import copy
import shutil
import openpyxl
import pytz

from scipy.sparse import csr_matrix
from scipy.sparse.csgraph import connected_components

import matplotlib.pyplot as plt
from config import config
from dictionary_driver import Dictionary_Driver
from dictionary_manager import Dictionary_Manager

class NumpyEncoder(json.JSONEncoder):
    """ Special json encoder for numpy types """
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        return json.JSONEncoder.default(self, obj)


def SaveJsonData(data, path):
        try:
                file = open(path, 'wt+')
                json.dump(data, file, cls = NumpyEncoder)
                file.close()
        except:
                raise Exception("Couldn't open/write/close file:" + path)


def LoadJsonData(path):
        data = None

        try:
                file = open(path, 'rt')
                data = json.load(file)
                file.close()
        except:
                #raise Exception("Couldn't open/read/close file:" + path)
                pass

        return data

def SaveBinaryData(data, path):
        try:
                file = open(path, 'wb+')
                pickle.dump(data, file, pickle.HIGHEST_PROTOCOL)
                file.close()
        except:
                raise Exception("Couldn't open/read/save/close file:" + path)

def LoadBinaryData(path):
        data = None
        try:
                file = open(path, 'rb')
                data = pickle.load(file)
                file.close()
        except:
                # raise Exception("Couldn't open/read/close file:" + path)
                pass

        return data

def SaveDataFrame_Excel(df, filePath, sheet_name='Sheet1'):
        try:
                with pd.ExcelWriter(filePath) as writer:
                        df.to_excel(writer)
        except:
                raise Exception("Couldn't save to Excel file")
        return

def DeleteFile( path ):
        os.remove( path )

def read_excel(filename, sheet_name=0):
        # print(filename)
        # excelPath = os.path.join(countryFolder, filename + '.xlsx')

        # converters takes effect only after the excel file is read and, 
        # so, only after date-like string is converted to date type.
        df = pd.read_excel(filename, sheet_name=sheet_name, parse_dates=False) #, converters={'Date':str})
        return df

def read_csv(filename):
        # print(filename)
        # excelPath = os.path.join(countryFolder, filename + '.xlsx')
        df = pd.read_csv(filename, encoding='unicode_escape', sep=',')
        return df

def read_large_excel(data_foloder, filename):
        df = None
        binPath = os.path.join(data_foloder, filename + '.bin')
        bin = LoadBinaryData(binPath)
        if bin is not None:
            df = bin
        else:
            excelPath = os.path.join(data_foloder, filename + '.xlsx')
            df = pd.read_excel(excelPath)
            SaveBinaryData(df, binPath)
        return df

scheme_contest_data = {
        # 12
        # The columns of the following column names are either found in the excel file or created by EnhanceReadExcelFileForArrays(.)

        #'past_data_path' : "contest_past_data_partial.xlsx", \
        'past_data_path' : "contest_train_and_valid.xlsx", \
        'past_data_bin_path' : "contest_train_and_valid.bin", \
        'future_data_path' : "contest_testing_data.xlsx", \
        'past_data_generated' : 'contest_past_data_generated.xlsx', \
        #--------------------------------------------------------------------------------------- database-specific columns
        'gameIdCol' : ["fixture_id"],
        'leagueCountryCol' : ["league_country"], # as is, 'Unknown' - filled \
        'homeCountryCol' : ["home_team_country"], # as is, 'Unknown' - filled \
        'awayCountryCol' : ["away_team_country"], # as is, 'Unknown' - filled \
        'outcomeCol' : ['outcome'], # as is, [0,1,2] -> [-1, 0, 1] \
        #--------------------------------------------------------------------------------------- standard columns
        'leagueCol' : ['league_id'], # as is, srting-ized \
        'homeCol' : ['teams_home_team_id'], # as is, srting-ized \
        'awayCol' : ['teams_away_team_id'], # as is, srting-ized \
        'oddCols' : [ 'winning_odds_home', 'winning_odds_draw', 'winning_odds_away' ], # generated \,\
        'goalCols' : ['goals_home', 'goals_away'], \
        
        'situationCols' : ['year', 'month', 'dayofweek', # generated \
                
                'goals_home', 'goals_away', # as is, mean - filled \

                'winning_percent_home',	'winning_percent_draws', 'winning_percent_away', \
                
                #'comparison_forme_home', 'comparison_forme_away', 'comparison_att_home', \
                #'comparison_att_away', 'comparison_def_home', 'comparison_def_away', 'comparison_fish_law_home', 'comparison_fish_law_away', 'comparison_h2h_home', \
                #'comparison_h2h_away', 'comparison_goals_h2h_home', 'comparison_goals_h2h_away', 'teams_home_last_5_matches_forme', 'teams_home_last_5_matches_att', \
                #'teams_home_last_5_matches_def', 'teams_away_last_5_matches_forme', 'teams_away_last_5_matches_att', 'teams_away_last_5_matches_def', \
                #'teams_home_last_5_matches_goals', 'teams_home_last_5_matches_goals_avg', 'teams_home_last_5_matches_goals_against', 'teams_home_last_5_matches_goals_against_avg', \
                #'teams_home_all_last_matches_matchs_matchsPlayed_home', 'teams_home_all_last_matches_matchs_matchsPlayed_away', 'teams_home_all_last_matches_matchs_wins_home', \
                #'teams_home_all_last_matches_matchs_wins_away', 'teams_home_all_last_matches_matchs_draws_home', 'teams_home_all_last_matches_matchs_draws_away', \
                #'teams_home_all_last_matches_matchs_loses_home', 'teams_home_all_last_matches_matchs_loses_away', 'teams_home_all_last_matches_goals_goalsFor_home', \
                #'teams_home_all_last_matches_goals_goalsFor_away', 'teams_home_all_last_matches_goals_goalsAgainst_home', 'teams_home_all_last_matches_goals_goalsAgainst_away', \
                #'teams_home_all_last_matches_goalsAvg_goalsFor_home', 'teams_home_all_last_matches_goalsAvg_goalsFor_away', 'teams_home_all_last_matches_goalsAvg_goalsAgainst_home', \
                #'teams_home_all_last_matches_goalsAvg_goalsAgainst_away', 'teams_home_last_h2h_played_home', 'teams_home_last_h2h_played_away', 'teams_home_last_h2h_wins_home', \
                #'teams_home_last_h2h_wins_away', 'teams_home_last_h2h_draws_home', 'teams_home_last_h2h_draws_away', 'teams_home_last_h2h_loses_home', 'teams_home_last_h2h_loses_away', \
                #'teams_away_last_5_matches_goals', 'teams_away_last_5_matches_goals_avg', 'teams_away_last_5_matches_goals_against', 'teams_away_last_5_matches_goals_against_avg', \
                #'teams_away_all_last_matches_matchs_matchsPlayed_home', 'teams_away_all_last_matches_matchs_matchsPlayed_away', 'teams_away_all_last_matches_matchs_wins_home', \
                #'teams_away_all_last_matches_matchs_wins_away', 'teams_away_all_last_matches_matchs_draws_home', 'teams_away_all_last_matches_matchs_draws_away', \
                #'teams_away_all_last_matches_matchs_loses_home', 'teams_away_all_last_matches_matchs_loses_away', 'teams_away_all_last_matches_goals_goalsFor_home', \
                #'teams_away_all_last_matches_goals_goalsFor_away', 'teams_away_all_last_matches_goals_goalsAgainst_home', 'teams_away_all_last_matches_goals_goalsAgainst_away', \
                #'teams_away_all_last_matches_goalsAvg_goalsFor_home', 'teams_away_all_last_matches_goalsAvg_goalsFor_away', 'teams_away_all_last_matches_goalsAvg_goalsAgainst_home', \
                #'teams_away_all_last_matches_goalsAvg_goalsAgainst_away', 'teams_away_last_h2h_played_home', 'teams_away_last_h2h_played_away', 'teams_away_last_h2h_wins_home', \
                #'teams_away_last_h2h_wins_away', 'teams_away_last_h2h_draws_home', 'teams_away_last_h2h_draws_away', 'teams_away_last_h2h_loses_home', 'teams_away_last_h2h_loses_away', \
        ], \
        #------------------------------------------------------------------------------------------------------------
        'maxOdds' : 50.0, # Games with more than MaxOdds are dropped from history data or replaced by MaxOddss. Remove compiled database upon change of this. \
        'treat_maxOdds' : 'replace', #'drop',  #'replace', # or 'drop' \
        'maxGoals' : 10, # Games with more than MaxGoals are dropped from history data or replaced by MaxGoals. Remove compiled database upon change of this. \
        'treat_maxGoals' : 'replace' # 'drop', # or 'replace' \
}

def standardize_contest_data(df):

        df.replace('', '')
        df.replace('N', '')
        df.replace('NA', '')

        if 'erows' not in df.columns:
            print("creating 'erows'...")
            erows = [ row for row in range( 2, df.shape[0] + 2 ) ] # 2, because Excel rows starts with 2 rather than 0.
            df.insert( 0, 'erows', erows )

        #m_date = super_config['.date_org'] + datetime.timedelta( days = super_config['.M.Day'] )
        #dataUse['C.Date'] = super_config['.date_org'] + datetime.timedelta( days = super_config['.C.Day'] )

        if 'date' not in df.columns:
            print("creating 'date'...")
            dates = [ datetime.date( year, month, day ) for year, month, day in \
                    zip( pd.DatetimeIndex( df['fixture_date'] ).year, pd.DatetimeIndex( df['fixture_date'] ).month, pd.DatetimeIndex( df['fixture_date'] ).day ) ]
            df.insert( 1, 'date', dates )

        df.sort_values( by = ['date'], ascending = False, inplace = True ) #----------------------------- Sort by Date.

        #==========  Reverse the row order.
        df = df.iloc[::-1]

        #========== Fill in empty country names.
        df.loc[ pd.isnull( df['league_country'] ), 'league_country' ] = 'Unknown'
        df.loc[ pd.isnull( df['home_team_country'] ), 'home_team_country' ] = 'Unknown'
        df.loc[ pd.isnull( df['away_team_country'] ), 'away_team_country' ] = 'Unknown'

        #========== [0,1,2] -> [-1,0,1]: 2->-1 ( home_win : 1, draw : 0, away_win : -1 )
        df.loc[ df['outcome'] == 2, 'outcome' ] = -1

        #========== String-ize league_id, home_id, and away_id
        df[ 'league_id' ] =  [ str( league_number ) for league_number in df[ 'league_id' ]  ]
        df[ 'teams_home_team_id' ] = [ str( home_number ) for home_number in df[ 'teams_home_team_id' ]  ]
        df[ 'teams_away_team_id' ] = [ str( away_number ) for away_number in df[ 'teams_away_team_id' ]  ]

        #========== Add odds for home, draw, and away.
        bfr = 0.05 # eg. 0.05. bfr == bookie_fee_rate. The percentage of ( 1/p -1 ), or the bookies share in the amount that the bettor actually creates.
        wp = df['winning_percent_home']; loc = df.columns.get_loc('winning_percent_home') + 1
        df.insert( loc, 'winning_odds_home', 100 / ( wp + 0.5 ) * ( 1.0 - bfr ) + bfr )
        wp = df['winning_percent_draws']; loc = df.columns.get_loc('winning_percent_draws') + 1
        df.insert( loc, 'winning_odds_draw', 100 / ( wp + 0.5 ) * ( 1.0 - bfr ) + bfr )
        wp = df['winning_percent_away']; loc = df.columns.get_loc('winning_percent_away') + 1
        df.insert( loc, 'winning_odds_away', 100 / ( wp + 0.5 ) * ( 1.0 - bfr ) + bfr )

        #========== Add year, month, and ndow.
        loc = df.columns.get_loc( 'fixture_date' )
        df.insert( loc + 1, 'year', pd.DatetimeIndex( df['fixture_date'] ).year - 2019  )
        df.insert( loc + 2, 'month', pd.DatetimeIndex( df['fixture_date'] ).month - 6 )
        df.insert( loc + 3, 'dayofweek', pd.DatetimeIndex( df['fixture_date'] ).dayofweek - 3 )

        #========== Fill in empty goals_*  NOTE goals_home/away are all negative.
        df.loc[ pd.isnull(df['goals_home']), 'goals_home' ] = df['goals_home'].mean()
        df.loc[ pd.isnull(df['goals_away']), 'goals_away' ] = df['goals_away'].mean()

        #========== outcome
        df.loc[ pd.isnull(df['outcome']), 'outcome' ] = 1


        nan_value = float( "NaN" )

        #------------------------------------------------------------------------ Regularize odds columns.
        if scheme_contest_data['treat_maxOdds'] == 'drop' :
                for oddCol in scheme_contest_data['oddCols']: df.loc[ df[oddCol] > scheme_contest_data['maxOdds'] , [oddCol] ] = nan_value
                df.dropna( axis = 0, subset = scheme_contest_data['oddCols'], inplace=True )
        elif scheme_contest_data['treat_maxOdds'] == 'replace' :
                for oddCol in scheme_contest_data['oddCols']: df.loc[ df[oddCol] > scheme_contest_data['maxOdds'] , [oddCol] ] = scheme_contest_data['maxOdds']
        else : raise Exception('Wrong data scheme.')

        # ------------------- Regularize goals for thomas database.
        if scheme_contest_data['treat_maxGoals'] == 'drop' :
                for goalCol in scheme_contest_data['goalCols']: df.loc[ df[goalCol] > scheme_contest_data['maxGoals'] , [goalCol] ] = nan_value
                df.dropna( axis = 0, subset = scheme_contest_data['goalCols'], inplace=True )
        elif scheme_contest_data['treat_maxGoals'] == 'replace' :
                for goalCol in scheme_contest_data['goalCols']: df.loc[ df[goalCol] > scheme_contest_data['maxGoals'] , [goalCol] ] = scheme_contest_data['maxGoals']
        else : raise Exception('Wrong data scheme.')

        #---------------------------------- Drop unused columns, and optionally save df for visual examination.
        columns_used = []
        for col in ['gameIdCol', 'leagueCol', 'homeCol', 'awayCol', 'oddCols', 'situationCols' ] : columns_used += scheme_contest_data[ col ]
        columns_used += ['erows', 'date']

        for col in [ 'leagueCountryCol', 'homeCountryCol', 'awayCountryCol', 'outcomeCol' ] : columns_used += scheme_contest_data[ col ]

        columns_unused = [ col for col in df.columns if col not in columns_used ]
        df.drop( axis = 1, columns = columns_unused ) # PENDING - They are not dropped and remain, when seved to excel.

        # df.to_excel( scheme_contest_data['past_data_generated'] ) # Save back to a file for visual exam. Takes too long. Mask it for the moment.

        return df

def get_odds(probability, bookie_profit_percent):
        odds = (1.0 - bookie_profit_percent/100) / (probability + 1e-9)
        return odds

def get_probability(odds, bookie_profit_percent):
        probability = (1.0 - bookie_profit_percent/100) / odds - 1e-9
        return probability

def improve_contest_data(df):

        df.replace('', '')
        df.replace('N', '')
        df.replace('NA', '')

        #========== Fill in empty country names.
        df.loc[ pd.isnull( df['league_country'] ), 'league_country' ] = config['unknown_token']
        df.loc[ pd.isnull( df['home_team_country'] ), 'home_team_country' ] = config['unknown_token']
        df.loc[ pd.isnull( df['away_team_country'] ), 'away_team_country' ] = config['unknown_token']

        #========== Add odds for home, draw, and away.
        bfr = 0.05 # eg. 0.05. bfr == bookie_fee_rate. The percentage of ( 1/p -1 ), or the bookies share in the amount that the bettor actually creates.
        wp = df['winning_percent_home']; loc = df.columns.get_loc('winning_percent_home') + 1
        df.insert( loc, 'winning_odds_home', get_odds(wp/100, config['bookie_profit_percent']) ) # 100 / ( wp + 0.5 ) * ( 1.0 - bfr ) + bfr )
        wp = df['winning_percent_draws']; loc = df.columns.get_loc('winning_percent_draws') + 1
        df.insert( loc, 'winning_odds_draw',  get_odds(wp/100, config['bookie_profit_percent']) )  # 100 / ( wp + 0.5 ) * ( 1.0 - bfr ) + bfr )
        wp = df['winning_percent_away']; loc = df.columns.get_loc('winning_percent_away') + 1
        df.insert( loc, 'winning_odds_away',  get_odds(wp/100, config['bookie_profit_percent']) )  # 100 / ( wp + 0.5 ) * ( 1.0 - bfr ) + bfr )

        #========== Fill in empty goals_*  NOTE goals_home/away are all negative.
        df.loc[ pd.isnull(df['goals_home']), 'goals_home' ] = df['goals_home'].mean()
        df.loc[ pd.isnull(df['goals_away']), 'goals_away' ] = df['goals_away'].mean()

        #========== outcome
        df.loc[ pd.isnull(df['outcome']), 'outcome' ] = 1


        nan_value = float( "NaN" )

        #------------------------------------------------------------------------ Regularize odds columns.
        if scheme_contest_data['treat_maxOdds'] == 'drop' :
                for oddCol in scheme_contest_data['oddCols']: df.loc[ df[oddCol] > scheme_contest_data['maxOdds'] , [oddCol] ] = nan_value
                df.dropna( axis = 0, subset = scheme_contest_data['oddCols'], inplace=True )
        elif scheme_contest_data['treat_maxOdds'] == 'replace' :
                for oddCol in scheme_contest_data['oddCols']: df.loc[ df[oddCol] > scheme_contest_data['maxOdds'] , [oddCol] ] = scheme_contest_data['maxOdds']
        else : raise Exception('Wrong data scheme.')

        # ------------------- Regularize goals for thomas database.
        if scheme_contest_data['treat_maxGoals'] == 'drop' :
                for goalCol in scheme_contest_data['goalCols']: df.loc[ df[goalCol] > scheme_contest_data['maxGoals'] , [goalCol] ] = nan_value
                df.dropna( axis = 0, subset = scheme_contest_data['goalCols'], inplace=True )
        elif scheme_contest_data['treat_maxGoals'] == 'replace' :
                for goalCol in scheme_contest_data['goalCols']: df.loc[ df[goalCol] > scheme_contest_data['maxGoals'] , [goalCol] ] = scheme_contest_data['maxGoals']
        else : raise Exception('Wrong data scheme.')

        return df

def array_dict_contest_date(df):
       
        gameIds = np.array( df[ scheme_contest_data['gameIdCol'] ], dtype = np.int64 ); gameIds = np.reshape( gameIds, (-1) )
        erows = np.array( df[ 'erows' ], dtype = np.int64 ); erows = np.reshape( erows, (-1) )
        dates = np.array( df[ 'date' ] ); dates = np.reshape( dates, (-1) )
        leagues = np.array( df[ scheme_contest_data['leagueCol'] ], dtype = np.str ); leagues = np.reshape( leagues, (-1) )
        homes = np.array( df[ scheme_contest_data['homeCol'] ], dtype = np.str ); homes = np.reshape(homes, (-1) )
        aways = np.array( df[ scheme_contest_data['awayCol'] ], dtype = np.str ); aways = np.reshape(aways, (-1) )
        odds = np.array( df[ scheme_contest_data['oddCols'] ], dtype = np.float32 ) # No: odds = np.reshape( odds, (-1) )
        situation = np.array( df[ scheme_contest_data['situationCols'] ], dtype = np.float32 ) # No: situation = np.reshape( situation, (-1) )


        lCountry = np.array(df[ scheme_contest_data['leagueCountryCol'] ], dtype = np.str ); lCountry = np.reshape( lCountry, (-1) )
        hCountry = np.array(df[ scheme_contest_data['homeCountryCol'] ], dtype = np.str ); hCountry = np.reshape( hCountry, (-1) )
        aCountry = np.array(df[ scheme_contest_data['awayCountryCol'] ], dtype = np.str ); aCountry = np.reshape( aCountry, (-1) )

        outcome = np.array( df[ scheme_contest_data['outcomeCol'] ], dtype = np.float32 ); outcome = np.reshape( outcome, (-1) )

        data = {'gameIds': gameIds, 'erows': erows, 'leagues': leagues, 'homes': homes, 'aways': aways, 'situation': situation, 'odds': odds, 'lCountries': lCountry, 'hCountries': hCountry, 'aCountries' : aCountry, 'outcome' : outcome }
        data[ 'featureNames' ] = [ 'gameIds', 'erows', 'leagues', 'homes', 'aways', 'situation', 'odds', 'lCountries', 'hCountries', 'aCountries', 'outcome' ]

        return data

def CorrectSpells( data, dictionaries ) :
        lenth = len( data['leagues'] )
        variations = dictionaries.get( 'name_variations', None )
        string_features = [ 'leagues', 'homes', 'aways', 'lCountries', 'hCountries', 'aCountries' ]
        for n in range( lenth ) :
                for feature in string_features :
                        data[feature][n] = Regularize( data[feature][n] )                      
                        if variations is not None :  data[feature][n] = Find_Formal_Name( data[feature][n], variations )
        return data

def Regularize( string ):
        string = string.upper()
        string = re.sub( '\s+', '', string )    # remove all white space(s)
        return string

def Find_Formal_Name( string, variations ) :
        for key, list_val in variations.items() :
                if string in list_val :
                        string = key
                        print("formal name '{}' found for '{}': ".format(key, string))
        return string

def load_dictionaries():
        dictionary = Dictionary_Manager('')
        dictionaries = dictionary.Load_Dictionaries( filename = None, dict_names = ['leagues', 'teams', 'countries'] )
        name_variations = dictionary.Load_Dictionaries( filename = "Manually_Built Dictionary.xlsx", dict_names = ['team_name_variations'], list_mode = True )
        name_variations_total = {}
        for dict_name, dict_body in name_variations.items() : name_variations_total.update( dict_body )
        dictionaries['name_variations'] = name_variations_total

        return dictionaries

def save_dictionaries(dictionaries_org):
        dictionary = Dictionary_Manager('')

        dictionaries = dictionaries_org.copy()
        name_variations = dictionaries.get('name_variations', None)
        if name_variations is not None : dictionaries.pop('name_variations') # name_variations must not changed by code.
        clusters_dict = dictionaries.get( 'clusters', None )
        if clusters_dict is not None : dictionaries.pop( 'clusters' )

        dictionary.Save_Dictionaries( filename = None, dictionaries = dictionaries, overwrite = True )

        if clusters_dict is not None :
            dictionary.Save_Dictionaries( filename = None, dictionaries = { 'clusters' : clusters_dict }, overwrite = True, list_mode = True )
       

def Tokenize_contest_data( data, dictionaries ) :

        #if dictionaries.get( 'leagues', None ) is None : 
        #        string_array = data['leagues']
        #        dictionaries['leagues'] = CreateStringDict( string_array )

        variations = dictionaries.get( 'name_variations', None )
       
        data['leagues'], dictionaries['leagues'] = Translate( data['leagues'], dictionaries['leagues'], variations )
        data['nLeagues'] = len( dictionaries['leagues'] ) # nLeagues will not change even if some leagues will be removed from data. league_ids will spread all over [0, nLeagues).
        data['leagues_dic'] = dictionaries['leagues']

        #if dictionaries.get( 'teams', None ) is None : 
        #        string_array = np.concatenate( [ data['homes'], data['aways'] ], axis = 0 )
        #        dictionaries['teams'] = CreateStringDict( string_array )
       
        data['homes'], dictionaries['teams'] = Translate( data['homes'], dictionaries['teams'], variations )
        data['aways'], dictionaries['teams'] = Translate( data['aways'], dictionaries['teams'], variations )
        data['nTeams'] = len( dictionaries['teams'] ) # nTeams will not change even if some teams will be removed from data. team_ids will spread all over [0, nTeams).
        data['teams_dic'] = dictionaries['teams']

        data['lCountries'], dictionaries['countries'] = Translate( data['lCountries'], dictionaries['countries'], variations )
        data['hCountries'], dictionaries['countries'] = Translate( data['hCountries'], dictionaries['countries'], variations )
        data['aCountries'], dictionaries['countries'] = Translate( data['aCountries'], dictionaries['countries'], variations )
        data['nCountries'] = len( dictionaries['countries'] ) # nCountries will not change even if some countries will be removed from data. country_ids will spread all over [0, nCountries).
        data['countries_dic'] = dictionaries['countries']

        return data, dictionaries

# 4.1
def Translate( string_array, existing_dict, variations ) :

        #--------------------------------------------------------------------------------------------------------------
        #
        #       Design principle : Preserve existing part of dictionary, and APPEND ONLY.
        #
        #       Do NOT allow a name to change its token from time to time. 
        #       Do not revoke tokens from removed names.
        #       Newly introduced names can only have a newly introduced token.


        index_max = -1
        for string, index in existing_dict.items() :
                if index_max < index : index_max = index
        assert index_max + 1 == len( existing_dict ) # Not very strict restrict. The order of indices is not checked. Note: python dictionary now preserves the order of elements.

        tokenized_list = []
        for string in string_array :
                string = Regularize( string )
                
                index = existing_dict.get( string, -1 )
                if index >= 0 :
                        pass
                else: 
                        index = len( existing_dict ) # takes time? but elegant :) There will be few times to do this.
                        if variations is not None : string = Find_Formal_Name( string, variations )
                        existing_dict[ string ] = index

                tokenized_list.append( index ) # PENDING or NOTE : a place for unknown strings.

        dictionary_updated = existing_dict

        return np.array( tokenized_list ), dictionary_updated

def AugmentData( data ) :
        return data


def Select_Major_Games( data, dictionaries ) :

        # Premise: data is arranged in ascending order of game ids.

        nGames = len( data['homes'] ); homes = data['homes']; aways = data['aways']; erows = data['erows']; leagues = data['leagues']

        major_teams_list, dictionaries = Get_Major_Clustered_Teams( nGames, homes, aways, erows, leagues, dictionaries )
        major_teams = set( major_teams_list )


        multi_game_teams = set( Get_Multi_Game_Teams( homes, aways, minGames = config['minAttends'] ) ) # Attend at least 'minGames' games, either as home or away.
        major_teams = major_teams.intersection( multi_game_teams )
        major_game_ids = [ gameId for gameId in range( len(homes) ) if ( homes[gameId] in major_teams ) and ( aways[gameId] in major_teams ) ]
        major_game_ids.sort()

        def Select_Major( feature ): data[ feature ] = data[ feature ][ major_game_ids ]

        for feature in data['featureNames'] : Select_Major( feature )
        
        return data, dictionaries

# 6.1
def Get_Major_Clustered_Teams( nGames, homes, aways, erows, leagues, dictionaries ) :
        nTeams = ( max( max( homes ), max( aways ) ) - min( min( homes ), min( aways ) ) ) + 1
        print("Should equal nTeams: ", nTeams)
        data = np.array( [ 1 ] * (1 * nGames)) # a game between 'home' and 'away' teams is a connection/edge between the nodes 'home' and 'away'       
        graph = csr_matrix( ( data, ( np.concatenate([homes]), np.concatenate([aways]) ) ), shape = ( nTeams, nTeams ) )
        graph = graph.toarray()
        print("max graph: {}".format(np.max(graph)))
        
        # graph[i, j] : the number of games between the i-th team as the home team and j-th team as the away team.
        # graph[j, i] : i-th team as the away team and j-th team as the home team.
        # So, graph now represents a directed graph.

        # the number of home games plus the numebr away games of the i-th team.
        attend = [np.sum(graph[i, :]) + np.sum(graph[:, i]) for i in range(graph.shape[0])]     
        attend = np.array(attend, dtype=np.int64)
        print("attend vector = ", attend)
        busyTeamIds = np.where(attend >= config['minAttends'])[0]
        print("attend >= {}: {}".format(config['minAttends'], busyTeamIds))
        
        nShow = 10
        # Find all connected components. 
        # directed == False: i-th team and j-th team is connected if and only if graph[i, j] + graph[j, i] > 0.
        # labels[i] is the component label where i-th team belongs.
        nClusters, clusterIdMap = connected_components( csgraph = csr_matrix(graph), directed = False, return_labels = True )
        print("nClusters: {}, clusterIdMap: {}".format(nClusters, clusterIdMap))

        # unique[i] is the i-th components' label.
        # counts[i] is the number of teams in the i-th component.
        clusterIds, clusterSizes = np.unique( list( clusterIdMap ), return_counts = True )
        print("{} clusterIds: {}, {} clusterSizes: {}, total clusterSizes: {}".format(nShow, clusterIds[:nShow], nShow, clusterSizes[:nShow], np.sum(clusterSizes)))
        
        descendingIdxOrder = np.argsort(-clusterSizes)
        clusterIds = clusterIds[descendingIdxOrder]
        clusterSizes = clusterSizes[descendingIdxOrder]
        print("{} major_clusterSizes: {}".format(nShow, clusterSizes[:nShow]))
        print("{} major_clusterIds: {}".format(nShow, clusterIds[:nShow]))

        clusters = []
        for cId in clusterIds:
                c = {}
                c.update({ 'cId': cId })
                cTeamIds = np.array(range(nTeams))[np.where(clusterIdMap==cId)]
                c.update({'teamIds': cTeamIds})
                cGames = [erows[i] for i in range(erows.shape[0]) if homes[i] in cTeamIds and aways[i] in cTeamIds ]
                c.update({'games': cGames})
                cLeagues = list(set([leagues[i] for i in range(leagues.shape[0]) if homes[i] in cTeamIds and aways[i] in cTeamIds]))
                c.update({'leagues': cLeagues})            
                clusters.append(c)
        print('1st major cluster nTeamIds:', len(clusters[0]['teamIds']))
        print("1st major cluster {} teamIds: {}".format(nShow, clusters[0]['teamIds'][:nShow]))
        print('1st major cluster nGames:', len(clusters[0]['games']))
        print("1st major cluster {} games: {}".format(nShow, clusters[0]['games'][:nShow]))
        print("1st major cluster nLeagues: {}".format(len(clusters[0]['leagues'])))
        print("1st major cluster all leagues: {}".format(clusters[0]['leagues']))
        for l in clusters[0]['leagues']:
               print(lookup(dictionaries['leagues'], l))

        leaguesList = []
        leagueIds = np.array(list(set(leagues)))
        lNGames = np.array([len([erows[i] for i in range(len(erows)) if leagues[i] == l]) for l in leagueIds])
        lNTeams = np.array([len(list(set([homes[i] for i in range(len(homes)) if leagues[i] == l]).union(set([aways[i] for i in range(len(aways)) if leagues[i] == l])))) for l in leagueIds])
        lGamesTeams = lNGames * lNTeams
        descendingIdxOrder = np.argsort(-lGamesTeams)
        leagueIds = leagueIds[descendingIdxOrder]

        for lId in leagueIds:
               l = {}
               l.update({'lId': lId})
               lGames = [erows[i] for i in range(erows.shape[0]) if leagues[i] == lId]
               l.update({'games': lGames})
               lTeamIds = list(set([homes[i] for i in range(nTeams) if leagues[i] == lId]).union(set([aways[i] for i in range(nTeams) if leagues[i] == lId])))
               l.update({'teamIds': lTeamIds})
               leaguesList.append(l)
        print('1st major league id: {}, league name: {}'.format(leaguesList[0]['lId'], lookup(dictionaries['leagues'], leaguesList[0]['lId'])))
        print('1st major league nTeamIds:', len(leaguesList[0]['teamIds']))
        print('1st major league nGame:', len(leaguesList[0]['games']))


        



        labels_of_major_clusters = clusterIds[np.where( clusterSizes >= config[ 'minClusterSize' ] )[0]]
        print("{} labels_of_major_clusters: {}".format(nShow, labels_of_major_clusters[:nShow]))





        indices = []
        if len( labels_of_major_clusters ) > 0 :
                indice_set = set()
                for n in range( len( labels_of_major_clusters) ) : 
                        indice_set = indice_set.union( set( np.where( clusterIdMap == labels_of_major_clusters[ n ] )[0] ) )
                indices = sorted( list ( indice_set ) )

        # labels[team_id] : the label of the component that the team of team_id belongs to.
        # dictionaries['teams'] = { team_name : team_id }
        unique_counts = np.stack( [ clusterIds, clusterSizes ], axis = 0 )
        unique_counts_sorted = unique_counts[ :, np.flip( unique_counts[ 1, : ].argsort() ) ] # sort all, including 'unique' component labels, in the ascending order of 'counts'.
       
        teams_dict = dictionaries['teams']
        component_dict = {}; cluster_id = -1
        for component_id_idx in range( len( unique_counts_sorted[ 0, : ] ) ) : #
                component_id = unique_counts_sorted[ 0, component_id_idx ].item()
                team_list = []
                for label_idx in range( len( clusterIdMap ) ) : # label_is serves as team_id. labels[team_id] : the label of the component that the team of team_id belongs to
                        if component_id == clusterIdMap[ label_idx ] :
                                for team_name, team_id in teams_dict.items() :
                                        if team_id == label_idx :
                                                team_list.append( team_name )
                cluster_id += 1
                component_dict[ str( cluster_id ) ] = team_list # Use cluster_id, instead of component_id, beacues component_dict would show items in key order.
        # dictionaries['clusters'] = component_dict

        return indices, dictionaries # ordered list of team ids.


# 6.2
def Get_Multi_Game_Teams( homes, aways, minGames ) : # Teams that attended no game or just a single game. They are of no help.

        unique, counts = np.unique( np.concatenate( [ homes, aways ], axis = -1 ), return_counts = True )
        multi_game_indices = np.where( counts >= minGames )
        multi_game_teams = unique[ multi_game_indices[0] ]

        multi_game_teams.sort()

        return multi_game_teams # ordered list of team ids. 

def lookup(dictionary, value):
       return list(dictionary.keys())[list(dictionary.values()).index(value)]

import networkx as nx

def createGameGraph(df, renew=True):
        graph = nx.Graph()
        for fixture_id, home, away, fixture_date in zip(list(df['fixture_id']), 
                                                        list(df['teams_home_team_id']), 
                                                        list(df['teams_away_team_id']),
                                                        list(df['fixture_date'])
                                                        ):
                if (home, away) not in graph.edges:
                        graph.add_edge(home, away)
                        graph[home][away]['games'] = []
                graph[home][away]['games'].append((fixture_id, fixture_date))

        # This block removes duplicate fixture_ids in the data, like 563602.
        # This block reduces number of games significantly, from 474000 to 3xxxxx. Comment out
        for edge in graph.edges:
                games = graph[edge[0]][edge[1]]['games']
                games = list(set(games))
                graph[edge[0]][edge[1]]['games'] = games

        return graph
                
def fixture_id_to_ids(
        df,
        gameGraph,
        a_conductance_search=0.1,
        b_conductance_search=1.5,
        conductance365=0.8, 
        conductanceMedTeam=0.2, 
        ids_max = 30,
        minConductanceStep = 0.002,
        renew=True,
        ):
               
        alpha = pow(conductance365, -1/365)
        minCon = 1e-9

        def createRegistanceView(gameGraph, baseId, baseHome, baseAway, baseDate, weight='resistance'):
                def conductance(_date):
                        return pow(alpha, (_date - baseDate).days)

                # calculate gameGraph[.][.]['resistance'] based on base_date
                # cnt_if = 0; cnt_else = 0
                for (teamA, teamB) in gameGraph.edges:  # generats unique edges: both (1, 2) and (2, 1) are represented as (1, 2) 
                        con = minCon
                        for (_id, _date) in gameGraph[teamA][teamB]['games']:
                                if(baseDate <= _date):
                                        # print('if', cnt_if, cnt_else, _id, baseId)
                                        con += 0.0
                                        # cnt_if += 1      
                                else:   # _id != baseId and _date <= baseDate
                                        # print('else', cnt_else, _id, baseId)
                                        if(teamA==baseHome and teamB==baseAway or teamB==baseHome and teamA==baseAway):
                                                # the game's two teams are baseHome and baseAway. The conductance is the standard.
                                                con += conductance(_date)     
                                        else:   
                                                # the game has a team that is neither baseHome nor baseAway. The conductance is less than the standard.
                                                con += 1 / (1 / conductanceMedTeam + 1 / conductance(_date) )
                                        # cnt_else += 1
                        gameGraph[teamA][teamB][weight] = 1/con
                        # print('con: ', con, 1/con, cnt_if, cnt_else)

                return gameGraph
        
        def getPossibleUniquePairsFromDijkstras(homeDijkstra, awayDijkstra):
                reachablesHome = list(set(list(homeDijkstra.keys())))
                pairsHome = [(teamA, teamB) for teamA in reachablesHome for teamB in reachablesHome if teamA < teamB]
                reachablesAway = list(set(list(awayDijkstra.keys())))
                pairsAway = [(teamA, teamB) for teamA in reachablesAway for teamB in reachablesAway if teamA < teamB]

                pairs = list(set(pairsHome + pairsAway))

                return pairs
        
        def getGames(df, pairs, baseId, baseDate):
                fixture_ids = []
                sub_df = None
                for (teamA, teamB) in pairs:
                        inc_df = df.loc[(((df['teams_home_team_id']==teamA) & (df['teams_away_team_id']==teamB)) 
                        | ((df['teams_home_team_id']==teamB) & (df['teams_away_team_id']==teamA))
                        ) & (df['fixture_id'] != baseId) & (df['fixture_date'] <= baseDate), ['fixture_id', 'fixture_date']]  # include (df['fixture_id'] != baseId) and include equal sign.
                        sub_df = pd.concat([sub_df, inc_df])

                        # inc_df = df.loc[(((df['teams_home_team_id']==teamA) & (df['teams_away_team_id']==teamB)) 
                        # | ((df['teams_home_team_id']==teamB) & (df['teams_away_team_id']==teamA))
                        # ), ['fixture_id', 'fixture_date']]  # include (df['fixture_id'] != baseId) and include equal sign.
                        # sub_df = pd.concat([sub_df, inc_df])

                if(sub_df is not None):
                        sub_df = sub_df.sort_values(by='fixture_date', ascending=False) ['fixture_id']
                        fixture_ids += list(sub_df)
                        # print('fixture_ids len org', len(fixture_ids))
                        fixture_ids = list(dict.fromkeys(fixture_ids))
                        # print('fixture_ids len', len(fixture_ids))

                return fixture_ids
        
        def getGamesFaster(gameGraph, possibleUniquePairs, baseId, baseDate):
                pairs = [(A, B) for (A, B) in possibleUniquePairs if (A, B) in gameGraph.edges()]
                dates_ids = []
                for (A, B) in pairs:
                        dates_ids += [(date, id) for (id, date) in gameGraph[A][B]['games'] if date < baseDate]
                # dates_ids = [((date, id) for (id, date) in gameGraph[A][B]['games'] if date <= baseDate and id != baseId) for (A, B) in pairs]
                # print(dates_ids)
                dates_ids.sort(reverse=True)
                ids_dates = [id for (ts, id) in dates_ids]
                # print(ids_dates)
                return ids_dates
        
        def getPair(df, baseId):
                frame = df.loc[df['fixture_id'] == baseId, ['teams_home_team_id', 'teams_away_team_id']]
                pairs = list(zip(list(frame['teams_home_team_id']), list(frame['teams_away_team_id'])))
                return pairs[0]
        
        def getGameSequence(gameGraph, baseEdge, baseId, baseDate, pathMinConductivity):
                # # I believe len(games), where game is the return of this function, is a monotonically decreasing function of pathMinConductivity.
                # gameGraph = createRegistanceView(gameGraph, baseId, baseEdge[0], baseEdge[1], baseDate, weight='resistance')
                homeDijkstra = nx.single_source_dijkstra_path(gameGraph, baseEdge[0], cutoff = 1/pathMinConductivity, weight='resistance')
                awayDijkstra = nx.single_source_dijkstra_path(gameGraph, baseEdge[1], cutoff = 1/pathMinConductivity, weight='resistance')
                # print('reachables: ', len(list(homeDijkstra.keys())), len(list(awayDijkstra.keys())))

                possibleUniquePairs = getPossibleUniquePairsFromDijkstras(homeDijkstra, awayDijkstra)
                games = getGamesFaster(gameGraph, possibleUniquePairs, baseId, baseDate)
                # games = getGames(df, possibleUniquePairs, baseId, baseDate)
                # print('ids: ', len(games))
                return games
        
        def sort_id_to_ids(id_to_ids):
                dates_ids = [(id_to_ids[id][0], id) for id in id_to_ids.keys()]
                dates_ids.sort()
                # print('dates_ids', dates_ids)
                id_to_ids = {str(id): id_to_ids[id][1] for (ts, id) in dates_ids}
                # print('ids: ', id_to_ids)
                return id_to_ids

                # frame = df.loc[df['fixture_id'] == baseId, ['teams_home_team_id', 'teams_away_team_id']]
                # pairs = list(zip(list(frame['teams_home_team_id']), list(frame['teams_away_team_id'])))
                # return pairs[0]

        count = 0
        fixture_id_to_ids = {} 
        for baseEdge in list(gameGraph.edges):  #-----------------------------------------------------------------------------
                baseGames = gameGraph[baseEdge[0]][baseEdge[1]]['games']
                # print('1. baseGames', baseGames)

                for (baseId, baseDate) in baseGames:        # required, as diffrent games have different dates leading to diffrernt conductivity
                        print('baseId numer, data point numer to build: ', count, len(list(fixture_id_to_ids.keys())), ', baseId, pair, baseDate: ', baseId, getPair(df, baseId), baseDate, end='\r')
                        count += 1

                        # if baseId != 576033: continue   #---------------------------------------

                        if fixture_id_to_ids.get(baseId, None) is not None:
                                pass
                                # print("baseEdge: {}, baseId: {}".format(baseEdge, baseId))
                                # raise("data error: duplicate texture_id found:")
                        else:
                                gameGraph = createRegistanceView(gameGraph, baseId, baseEdge[0], baseEdge[1], baseDate, weight='resistance')
                                # Although know max conductivity of a single game is 1.0, we start from b_conductance_search of more than 1.0
                                # bacause multiple games between a home and a away might increase conductance between them to more than 1.0.
                                a = a_conductance_search; b = b_conductance_search
                                games_a = getGameSequence(gameGraph, baseEdge, baseId, baseDate, a); ids_a = len(games_a)
                                # print(games_a)
                                games_b = getGameSequence(gameGraph, baseEdge, baseId, baseDate, b); ids_b = len(games_b)
                                # print(games_b)

                                while True:
                                        # print(ids_a, ids_b)
                                        if ids_a <= ids_max or ids_a == ids_b:  # no hope
                                                games = games_a
                                                break
                                        elif ids_b >= ids_max:  # no hope
                                                games = games_b
                                                break
                                        else: # find optimal pathMinConductivity in (a, b)
                                                c = (a + b) / 2
                                                # print('c = ', c)
                                                if c - a <= minConductanceStep:
                                                        # if ids_a < ids_b:
                                                        #         games = games_b
                                                        # else:
                                                        #         games = games_a

                                                        if abs(ids_max-ids_a) <= abs(ids_max-ids_b):
                                                                games = games_a
                                                        else:
                                                                games = games_b
                                                        break
                                                else:   # find in (a, b). we know ids_a > ids_max, idx_a != ids_b, ids_max > ids_b
                                                        games_c = getGameSequence(gameGraph, baseEdge, baseId, baseDate, c)
                                                        ids_c = len(games_c)
                                                        # print(ids_a, ids_b, ids_c)
                                                        if ids_a > ids_max and ids_max > ids_c: # we have a hope
                                                                b = c; ids_b = ids_c; games_b = games_c
                                                        elif ids_c >= ids_max and ids_max > ids_b:      # we have a hope
                                                                a = c; ids_a = ids_c; games_a = games_c
                                                        else:   # ids_max > idx_b
                                                                print('logical error')
                                                                # raise("logical error")
                                
                                                             
                                # games = getGameSequence(gameGraph, baseEdge, baseId, baseDate, pathMinConductivity)

                                if games is not None and len(games) >= 0:       # include no past-sequence games. 
                                        fixture_id_to_ids.update({baseId: (baseDate, games)})

        fixture_id_to_ids = sort_id_to_ids(fixture_id_to_ids)


        return fixture_id_to_ids

def dataPrams_to_filename(
        bookie_profit_percent, a_conductance_search, b_conductance_search, conductance365, 
        conductanceMedTeam, ids_max, minConductanceStep, hours, nDataPoints, maxLen, prefix = 'ids'
):
    filename = "-".join([prefix, str(bookie_profit_percent), str(a_conductance_search), str(b_conductance_search), str(conductance365), \
                        str(conductanceMedTeam), str(ids_max), str(minConductanceStep), str(hours), str(nDataPoints), str(maxLen)])
    return filename
def filename_to_dataParams(filename):
    (prefix, bookie_profit_percent, a_conductance_search, b_conductance_search, conductance365,
     conductanceMedTeam, ids_max, minConductanceStep, hours, nDataPoints, maxLen) = \
     filename.split('-')
    return (float(bookie_profit_percent), float(a_conductance_search), float(b_conductance_search), float(conductance365),
     float(conductanceMedTeam), int(ids_max), float(minConductanceStep), float(hours), int(nDataPoints), int(maxLen))

def get_master_df_from_football_data_co_uk(countryFolder):
        master = None

        binPath = os.path.join(countryFolder, "master_improved" + '.bin')
        excelPath = os.path.join(countryFolder, "master_improved" + '.xlsx')
        df = LoadBinaryData(binPath)

        if df is not None:
                master = df
        else:
                csvFiles = []
                for (dirpath, dirnames, files) in os.walk(countryFolder):     # Make sure the .csv files are renamed .xlsx.
                        for filename in files:
                                if '.xlsx' in filename and '~' not in filename:
                                        csvFiles.append((dirpath, filename))
                # print(len(csvFiles)) 

                dframes = []
                for (countryFolder, csvFilename) in csvFiles:
                        filePath = os.path.join(countryFolder, csvFilename)
                        # print('path', countryFolder, filePath)
                        df = read_excel(filePath)
                        dframes.append(df)
                # print(len(dframes))

                common_cols = dframes[0].columns
                for df in dframes:
                        cols = df.columns
                        common_cols = [col for col in cols if col in common_cols]
                        # print(len(common_cols), len(cols))
                # print(common_cols)

                master = None
                for df in dframes:
                        df = df[common_cols]
                        master = pd.concat([master, df])
                        # print(len(master.index), len(df.index), len(master.columns))

                master = improve_uk_dataframe(master)

                SaveBinaryData(master, binPath)
                SaveDataFrame_Excel(master, excelPath)

        return master


def get_grown_and_new_from_football_data(countryFolder, Non_Odds_cols, num_bookies=5, oddsGroupsToExclude = [], train_mode=True, skip=False):
        binPath_grown = os.path.join(countryFolder, "df_grown" + '.bin')
        excelPath_grown = os.path.join(countryFolder, "df_grown" + '.xlsx')
        binPath_new = os.path.join(countryFolder, "df_new" + '.bin')
        excelPath_new = os.path.join(countryFolder, "df_new" + '.xlsx')
        binPath_white = os.path.join(countryFolder, "df_white" + '.bin')
        excelPath_white = os.path.join(countryFolder, "df_white" + '.xlsx')
        dictPath = os.path.join(countryFolder, "df_grown" + '.json')

        if not train_mode or skip:      # Try to use the existing ones if in inference mode or requested to skip.
                df_grown = LoadBinaryData(binPath_grown)
                df_new = LoadBinaryData(binPath_new)
                if df_grown is not None and df_new is not None:
                        return df_grown, df_new

        #-------------- collect all .xlsx files in the target foloder.
        csvFiles = []
        for (dirpath, dirnames, files) in os.walk(countryFolder):     # Make sure the .csv files are renamed .xlsx.
                for filename in files:
                        # foloders (dirpath) that ends with '_' are skipped.
                        if not dirpath.endswith('_') and '.xlsx' in filename and '~' not in filename and 'df_' not in filename:
                                print(filename, end=' ')
                                csvFiles.append((dirpath, filename))

        csvFiles.sort()
        print(len(csvFiles), ' csv files found') 

        #-------------- find odds groups that have 1X2 odds
        def find_1X2_odds_group(columns):
                oGroups = []
                i = 0
                while i <= len(columns)-3:
                        if columns[i].endswith('H') and columns[i+1].endswith('D') and columns[i+2].endswith('A') \
                        and columns[i][:-1] == columns[i+1][:-1] and columns[i+1][:-1] == columns[i+2][:-1]:
                                oGroups.append(columns[i][:-1]); i += 3
                        else: i += 1
                return oGroups

        #-------------- read all excel files, and find odds groups, and the total of them, that have 1X2 odds.
        oGroups = []
        filename_dict = {}
        for (_dirPath, csvFilename) in csvFiles:
                filePath = os.path.join(_dirPath, csvFilename)
                # df = read_csv(filePath) doesn't work with football-data.co.uk data.
                df = read_excel(filePath)
                # print('-------------- time :',  df['Time'][0])
                columns = df.columns

                filesToSkip = []
                if set(Non_Odds_cols) <= set(columns): pass
                else:
                        print(filePath + " will be dropped out because it lacks some of the required columns !!!")
                        filesToSkip.append(csvFilename)
                
                # print(columns)
                ogs = find_1X2_odds_group(columns)
                # print(csvFilename, ogs)
                non_ogs = [col for col in columns if col[:-1] not in ogs]
                # print(non_ogs)
                filename_dict[csvFilename] = (df, non_ogs, ogs)
                # print(ogs)
                oGroups += ogs

        #-------------- Remove fils that lack some of required non-odds columns.
        for file in filesToSkip:
                filename_dict.pop(file, None)

        unique_oGroups = list(set(oGroups))
        # print(unique_oGroups)
        # print('unique oGroups: ', unique_oGroups)
        counts = [(oGroups.count(og), og) for og in unique_oGroups]
        counts.sort(reverse=True)
        # print('oGroup counts: ', counts)
        if len(counts) < num_bookies:
                print("!!!!!! Too many bookies requested.")
        total_ogs_ordered = [og for (_, og) in counts][: min(num_bookies, len(counts))]     #-------- limited to the first num_bookies bookies.
        total_ogs_ordered = [og for og in total_ogs_ordered if og not in oddsGroupsToExclude]   # ---- this may reduce the list.

        def get_1X2_rename_dict(local_ogs):
                local_ogs_ordered = [og for og in total_ogs_ordered if og in local_ogs]
                rename_plan = { local_ogs_ordered[i] : 'HDA' + str(i) for i in range(len(local_ogs_ordered)) }
                rename_dict = {}
                for (source_og, target_og) in rename_plan.items():
                        rename_dict = rename_dict | \
                        { source_og + 'H' : target_og + 'H', source_og + 'D' : target_og + 'D', source_og + 'A' : target_og + 'A'}
                return rename_plan, rename_dict

        # Choose an initial common_cols
        common_cols = None
        for (filename, (df, non_ogs, ogs)) in filename_dict.items():
                rename_plan, rename_dict = get_1X2_rename_dict(ogs)
                df.rename(columns=rename_dict, inplace=True)
                common_cols = df.columns
                common_cols = [col for col in common_cols if col[:-1] not in oddsGroupsToExclude]
                break

        # Refine common_cols
        for (filename, (df, non_ogs, ogs)) in filename_dict.items():
                rename_plan, rename_dict = get_1X2_rename_dict(ogs)
                df.rename(columns=rename_dict, inplace=True)
                filename_dict[filename] = (df, non_ogs, ogs, rename_plan)        # keep rename_plan for later later maintainance use.
                cols = df.columns
                common_cols = [col for col in cols if col in common_cols]

        # print('common cols: ', common_cols)
        assert set(Non_Odds_cols) <= set(common_cols)

        df_built = None
        for (filename, (df, non_ogs, ogs, rename_plan)) in filename_dict.items():
                if df_built is None:
                        df_built = df[common_cols]
                else:
                        df_built = pd.concat([df_built, df[common_cols]])
                # print('df_built rows: ', df_built.shape[0], list(df_built.columns))
                # filename_dict[filename] = (non_ogs, ogs, rename_plan)
                filename_dict[filename] = (rename_plan)

        # Display rename vector, for how inconsistent the bookie rename is.
        for i in range(num_bookies):
                key = 'HDA' + str(i)
                renameVector = []
                for (filename, (rename_plan)) in filename_dict.items():
                        for (org_name, new_name) in rename_plan.items():
                                if key == new_name:
                                        renameVector.append(org_name)
                                        break
        
        # df_built.drop_duplicates()
        # Let this be the only place where df is sorted.
        df_built = df_built.sort_values(['Date', 'Div', 'HomeTeam', 'AwayTeam'], ascending=[True, True, True, True])   # This is required for consistency between calls.
        # No rows were dropped upto this point. Now permanent unique 'id's are assigned to the row.
        # A row's id assigned here will not change lifetime, UNLESS the same row was already assigned earlier.
        # 'id's will be used in place of ('Date', 'HomeTeam', 'AwayTeam'), for shorter storage space.
        baseId = config['baseGameId'] + 1   #--------------------------------------------------------------------
        ids = range(baseId, baseId+df_built.shape[0])
        df_built.insert(loc=0, column='id', value=ids)

        df_built = improve_uk_dataframe(df_built)

        df_grown_old = None
        df_grown_old = LoadBinaryData(binPath_grown)

        df_grown = df_new = None
        if df_grown_old is not None:
                pass # Make sure the rename_dictionaries are consistent.
                successful, df_grown, df_new = find_and_grow_with_extra_game_records_pythonic(df_grown_old, df_built)
                if successful:
                        print("Successfully found new game records and grew with them", "Extra rows: ", df_new.shape[0])

        if df_grown_old is not None:
                if successful:
                        if train_mode:
                                # df_grown is the total and saved, df_new is the tail part of df_total and saved into the new file.
                                if os.path.exists(binPath_grown): shutil.move(binPath_grown, os.path.join(countryFolder, "_Old_masters", "df_grown" + '.bin'))
                                if os.path.exists(excelPath_grown): shutil.move(excelPath_grown, os.path.join(countryFolder, "_Old_masters", "df_grown" + '.xlsx'))
                                if os.path.exists(binPath_new): shutil.move(binPath_new, os.path.join(countryFolder, "_Old_masters", "df_new" + '.bin'))
                                if os.path.exists(excelPath_new): shutil.move(excelPath_new, os.path.join(countryFolder, "_Old_masters", "df_new" + '.xlsx'))
                                if os.path.exists(dictPath): shutil.move(dictPath, os.path.join(countryFolder, "_Old_masters", "df_grown_x" + '.json'))
                                SaveBinaryData(df_grown, binPath_grown)
                                SaveDataFrame_Excel(df_grown, excelPath_grown)
                                SaveBinaryData(df_new, binPath_new)
                                SaveDataFrame_Excel(df_new, excelPath_new)
                                SaveJsonData(filename_dict, dictPath)
                        else:
                                # df_grown is the total and NOT saved, df_new is the tail part of df_total but saved into the WHITE file.
                                if os.path.exists(binPath_white): shutil.move(binPath_white, os.path.join(countryFolder, "_Old_masters", "df_white" + '.bin'))
                                if os.path.exists(excelPath_white): shutil.move(excelPath_white, os.path.join(countryFolder, "_Old_masters", "df_white" + '.xlsx'))
                                df_grown = df_grown_old         # 
                                SaveBinaryData(df_new, binPath_white)
                                SaveDataFrame_Excel(df_new, excelPath_white)  
                             
                else:
                        raise("!!! Failed to find and grow with extra game records")
        elif df_grown_old is None:      # successful is None
                if train_mode:
                        # df_grown is the total and saved, df_new is empty and saved into the new file.
                        df_grown = df_built
                        SaveBinaryData(df_grown, binPath_grown)         # df_built
                        SaveDataFrame_Excel(df_grown, excelPath_grown)  # df_built
                        SaveJsonData(filename_dict, dictPath)
                        # create an empty df_new
                        print("creating an empty df_new...")
                        df_new = df_grown.iloc[:0,:].copy()
                        SaveBinaryData(df_new, binPath_new)
                        SaveDataFrame_Excel(df_new, excelPath_new)
                else:
                        raise("!!! Inference mode requires df_grown to exist.")
                
        return df_grown, df_new


#-------------- find odds groups that have 1X2 odds
def find_1X2_odds_group(columns):
        oGroups = []
        i = 0
        while i <= len(columns)-3:
                if columns[i].endswith('H') and columns[i+1].endswith('D') and columns[i+2].endswith('A') \
                and columns[i][:-1] == columns[i+1][:-1] and columns[i+1][:-1] == columns[i+2][:-1]:
                        oGroups.append(columns[i][:-1]); i += 3
                else: i += 1
        return oGroups

def assign_seasonal_filenames(countryFolder):
    xlsxFiles = []
    for (dirpath, dirnames, files) in os.walk(countryFolder):     # Make sure the .csv files are renamed .xlsx.
        for filename in files:
            if '.xlsx' in filename and '~' not in filename and '-' not in filename and 'df_' not in filename:
                xlsxFiles.append((dirpath, filename))
    print('files found to rename: ', len(xlsxFiles))

    total_rows = 0
    for (_dirPath, xlsxFilename) in xlsxFiles:
            filePath = os.path.join(_dirPath, xlsxFilename)
            # df = read_csv(filePath) doesn't work with football-data.co.uk data.
            df = read_excel(filePath)
            org = df.shape[0]
            df = improve_uk_dataframe(df) #, dropNa=False)
            newFilePath = os.path.join(_dirPath, os.path.splitext(xlsxFilename)[0].split(' ')[0] + '-' + str(min(df['Date']).year) + '-' + str(max(df['Date']).year)+'.xlsx')
            # print(newFilePath)
            os.rename(filePath, newFilePath)
            total_rows += df.shape[0]
            del df
    print('total rows in renamed files: ', total_rows)
    return

def get_teams_by_div_from_total(countryTheme_folder_path):
    binPath_train = os.path.join(countryTheme_folder_path, "df_train" + '.bin')
    binPath_new = os.path.join(countryTheme_folder_path, "df_new" + '.bin')
    df_train = LoadBinaryData(binPath_train)
    df_new = LoadBinaryData(binPath_new)

    div_list = list(df_train['Div']); home_list = list(df_train['HomeTeam']); away_list = list(df_train['AwayTeam'])
    unique_divs = set(div_list)

    teams_by_div = { div : set([home_list[id] for id in range(len(home_list)) if div_list[id]==div]).union(set([away_list[id] for id in range(len(away_list)) if div_list[id]==div])) for div in unique_divs }
    for div1 in unique_divs:
        for div2 in unique_divs:
            if div1 != div2:
                assert teams_by_div[div1].intersection(teams_by_div[div2]) == set()
    teams_by_div = { div: list(teams) for (div, teams) in teams_by_div.items()}

    return teams_by_div


def CREATE_DADAFRAMES_v2(gameHistoryFolderPath, countryThemeFolder, Non_Odds_cols, num_bookies=5, oddsGroupsToExclude = [], preferedOrder = [], train_mode=True, skip=False):
        binPath_train = os.path.join(countryThemeFolder, "df_train" + '.bin')
        excelPath_train = os.path.join(countryThemeFolder, "df_train" + '.xlsx')
        binPath_new = os.path.join(countryThemeFolder, "df_new" + '.bin')
        excelPath_new = os.path.join(countryThemeFolder, "df_new" + '.xlsx')
        binPath_white = os.path.join(countryThemeFolder, "df_white" + '.bin')
        excelPath_white = os.path.join(countryThemeFolder, "df_white" + '.xlsx')
        dictPath = os.path.join(countryThemeFolder, "df_train" + '.json')

        if skip and train_mode:      # Try to use the existing ones if in inference mode or requested to skip.
                df_train = LoadBinaryData(binPath_train)
                df_new = LoadBinaryData(binPath_new)
                if df_train is not None and df_new is not None:
                        return df_train, df_new

        #-------------- collect all .xlsx files in the target foloder.
        csvFiles = []
        for (dirpath, dirnames, files) in os.walk(gameHistoryFolderPath):     # Make sure the .csv files are renamed .xlsx.
                for filename in files:
                        # foloders (dirpath) that ends with '_' are skipped.
                        if not dirpath.endswith('_') and '.xlsx' in filename and '~' not in filename and 'df_' not in filename:
                                print(filename, end=' ')
                                csvFiles.append((dirpath, filename))
        print()
        csvFiles.sort()
        print(len(csvFiles), ' files found') 


        #-------------- read all excel files, and find odds groups, and the total of them, that have 1X2 odds.
        oGroups = []
        filename_dict = {}
        for (_dirPath, csvFilename) in csvFiles:
                filePath = os.path.join(_dirPath, csvFilename)
                # df = read_csv(filePath) doesn't work with football-data.co.uk data.
                df = read_excel(filePath)
                # print('-------------- time :',  df['Time'][0])
                columns = df.columns

                filesToSkip = []
                if set(Non_Odds_cols) <= set(columns): pass
                else:
                        print(filePath + " will be dropped out because it lacks some of the required columns !!!")
                        filesToSkip.append(csvFilename)
                # print(columns)
                ogs = find_1X2_odds_group(columns)
                # print(csvFilename, ogs)
                non_ogs = [col for col in columns if col[:-1] not in ogs]

                # drop  oddsGroupsToExclude        
                for og in oddsGroupsToExclude:
                        try:
                                df.drop(og+'H', axis=1)
                                df.drop(og+'D', axis=1)
                                df.drop(og+'A', axis=1)
                        except: pass

                filename_dict[csvFilename] = (df, non_ogs, ogs)
                # print(ogs)
                oGroups += ogs

        #-------------- Remove fils that lack some of required non-odds columns.
        for file in filesToSkip:
                filename_dict.pop(file, None)

        unique_oGroups = list(set(oGroups))
        # print(unique_oGroups)
        # print('unique oGroups: ', unique_oGroups)
        counts = [(oGroups.count(og), og) for og in unique_oGroups]
        counts.sort(reverse=True)
        # print('oGroup counts: ', counts)
        if len(counts) < num_bookies:   print("!!!!!! Too many bookies requested.")

        total_ogs_ordered = [og for (_, og) in counts]

        # Execute preferedOrder, ignoring count order.
        indices = [i for i in range(len(total_ogs_ordered)) if total_ogs_ordered[i] in preferedOrder]
        for i in range(len(total_ogs_ordered)):
                if i in indices: total_ogs_ordered[i] = preferedOrder.pop(0)

        standard_ogs_ordered = total_ogs_ordered[:num_bookies]
        backup_ogs_ordered = total_ogs_ordered[num_bookies:]

        def get_1X2_rename_dict(local_ogs):
                # local_ogs_ordered = [og for og in total_ogs_ordered if og in local_ogs] #------- This will shift. We need a better solution.
                # assert len(local_ogs_ordered) >= num_bookies

                # if logal_ods misses any of standard_ogs_ordered, we borrow it.
                assert len(local_ogs) >= num_bookies
                local_backup_ordered = [og for og in backup_ogs_ordered if og in local_ogs]
                local_ogs_ordered = copy.deepcopy(standard_ogs_ordered)
                for id in range(len(local_ogs_ordered)):
                        if local_ogs_ordered[id] not in local_ogs:
                                local_ogs_ordered[id] = local_backup_ordered.pop(0)


                rename_plan = { local_ogs_ordered[i] : 'HDA' + str(i) for i in range(num_bookies) }
                rename_dict = {}
                for (source_og, target_og) in rename_plan.items():
                        rename_dict = rename_dict | \
                        { source_og + 'H' : target_og + 'H', source_og + 'D' : target_og + 'D', source_og + 'A' : target_og + 'A'}
                return rename_plan, rename_dict

        # Choose an initial common_cols
        common_cols = None
        for (filename, (df, non_ogs, ogs)) in filename_dict.items():
                rename_plan, rename_dict = get_1X2_rename_dict(ogs)
                df.rename(columns=rename_dict, inplace=True)
                common_cols = df.columns
                common_cols = [col for col in common_cols if col[:-1] not in oddsGroupsToExclude]
                break

        # Refine common_cols
        for (filename, (df, non_ogs, ogs)) in filename_dict.items():
                rename_plan, rename_dict = get_1X2_rename_dict(ogs)
                df.rename(columns=rename_dict, inplace=True)
                filename_dict[filename] = (df, non_ogs, ogs, rename_plan)        # keep rename_plan for later later maintainance use.
                cols = df.columns
                common_cols = [col for col in cols if col in common_cols]

        # print('common cols: ', common_cols)
        assert set(Non_Odds_cols) <= set(common_cols)

        df_built = None
        for (filename, (df, non_ogs, ogs, rename_plan)) in filename_dict.items():
                if df_built is None:
                        df_built = df[common_cols]
                else:
                        df_built = pd.concat([df_built, df[common_cols]])
                # print('df_built rows: ', df_built.shape[0], list(df_built.columns))
                # filename_dict[filename] = (non_ogs, ogs, rename_plan)
                filename_dict[filename] = (rename_plan)

        # Display rename vector, for how inconsistent the bookie rename is.
        for i in range(num_bookies):
                key = 'HDA' + str(i)
                renameVector = []
                for (filename, (rename_plan)) in filename_dict.items():
                        for (org_name, new_name) in rename_plan.items():
                                if key == new_name:
                                        renameVector.append(org_name)
                                        break
        
        # df_built.drop_duplicates()
        # Let this be the only place where df is sorted.
        # df_built = df_built.sort_values(['Date', 'Div', 'HomeTeam', 'AwayTeam'], ascending=[True, True, True, True])   # This is required for consistency between calls.
        df_built = df_built.sort_values(['Date', 'Div'], ascending=[True, True])   # This is required for consistency between calls.
        
        #=======================================================================================================
        # No rows were dropped upto this point. Now permanent unique 'id's are assigned to the row.
        # A row's id assigned here will not change lifetime, UNLESS the same row was already assigned earlier.
        # 'id's will be used in place of ('Date', 'HomeTeam', 'AwayTeam'), for shorter storage space.
        baseId = config['baseGameId'] + 1   #--------------------------------------------------------------------
        ids = range(baseId, baseId+df_built.shape[0])
        df_built.insert(loc=0, column='id', value=ids)

        df_built = improve_uk_dataframe(df_built, dropNa=False)

        df_train_old = None
        df_train_old = LoadBinaryData(binPath_train)

        df_train = df_new = None

        if df_train_old is not None:
                successful, df_train, df_new = find_and_grow_with_extra_game_records_pythonic(df_train_old, df_built)

                if successful:
                        print('train, new: ', df_train.shape, df_new.shape)
                        print("Successfully found new game records and grew with them", "New rows: ", df_new.shape[0])

                        if train_mode:
                                # df_train is the total and saved, df_new is the tail part of df_total and saved into the new file.
                                SaveBinaryData(df_train, binPath_train)
                                SaveDataFrame_Excel(df_train, excelPath_train)
                                SaveBinaryData(df_new, binPath_new)
                                SaveDataFrame_Excel(df_new, excelPath_new)
                                SaveJsonData(filename_dict, dictPath)
                        else:
                                # df_train is the total and NOT saved, df_new is the tail part of df_total but saved into the WHITE file.
                                df_train = df_train_old         # 
                                SaveBinaryData(df_new, binPath_white)
                                SaveDataFrame_Excel(df_new, excelPath_white)  
                             
                else:
                        raise("!!! Failed to find and grow with extra game records")
        elif df_train_old is None:      # successful is None
                if train_mode:
                        # df_train is the total and saved, df_new is empty and saved into the new file.
                        df_train = df_built
                        SaveBinaryData(df_train, binPath_train)         # df_built
                        SaveDataFrame_Excel(df_train, excelPath_train)  # df_built
                        SaveJsonData(filename_dict, dictPath)
                        # create an empty df_new
                        print("creating an empty df_new...")
                        df_new = df_train.iloc[:0,:].copy()
                        SaveBinaryData(df_new, binPath_new)
                        SaveDataFrame_Excel(df_new, excelPath_new)
                else:
                        raise("!!! Inference mode requires df_train to exist.")
                
        return df_train, df_new



def find_and_grow_with_extra_game_records(df, df_new):

    successful = True
    for index, row in df.iterrows():

        df_in_df = df[(df['home']==row['home']) & (df['away']==row['away']) & (df['date']==row['date'])]
        if df_in_df.shape[0] > 1:
            print('Found a df row duplicate in df.', row['home'], row['away'], row['date'])
            successful = False
            break
        
        df_in_df_new = df_new[(df_new['home']==row['home']) & (df_new['away']==row['away']) & (df_new['date']==row['date'])]
        if df_in_df_new.shape[0] < 1:
            print('Found a df row not in df_new.', row['home'], row['away'], row['date'])
            successful = False
            break
        elif df_in_df_new.shape[0] == 1:    # as expected !!!
            pass
            # print('Found a df row in df_new')
        else:
            print('Found a df row duplicate in df_new.', row['home'], row['away'], row['date'])
            successful = False
            break

    df_extra = None
    if successful:  # now, each df df rows is not duplicate and shows up in df_new just one time.
        for index, row in df_new.iterrows():
                print(row['home'], row['away'], row['date'])

                df_new_in_df = df[(df['home']==row['home']) & (df['away']==row['away']) & (df['date']==row['date'])]
                if df_new_in_df.shape[0] < 1:   # Extra row
                    # print('Found a df_new row not in df')

                    df_new_in_df_new = df_new[(df_new['home']==row['home']) & (df_new['away']==row['away']) & (df_new['date']==row['date'])]
                    if df_new_in_df_new.shape[0] > 1:
                        print('Found a df_new row duplicate in df_new.', row['home'], row['away'], row['date'])
                        successful = False
                        break
                    df_new_row = df_new[(df_new['home']==row['home']) & (df_new['away']==row['away']) & (df_new['date']==row['date'])]
                    if df_extra is None:
                        df_extra = pd.DataFrame(df_new_row)
                    else:
                        df_extra = pd.concat([df_extra, pd.DataFrame(df_new_row)])

    if successful: # now, each df rows is not duplicate and shows up in df_new just one time, plus, each of df_extra rows is not duplicate
            if df_extra is not None:
                id_list = list(range(max(df['id'])+1, max(df['id'])+1+df_extra.shape[0]))
                df_extra['id'] = id_list
                df_grown = pd.concat([df, df_extra])

    return successful, df_grown, df_extra


def find_and_grow_with_extra_game_records_pythonic(df, df_new):
    successful = True; df_grown = None; df_extra = None

    #====================================================================================
    # Totally, we must keep the id values of df rows preserved in the resulting df_grown
    #------------------------------------------------------------------------------------

    try:
        org_ids = list(df['id'])
        df = df.drop(['id'], axis=1)
        df_new = df_new.drop(['id'], axis=1)

        # Check if there are duplicate rows. -----------------------------------------
        # Do not drop but just check duplicate, because the org_id is there.
        assert (df.duplicated(subset=['Date', 'HomeTeam', 'AwayTeam']) == False).all()
        assert (df_new.duplicated(subset=['Date', 'HomeTeam', 'AwayTeam']) == False).all()

        # Check if df is a subset of df_new. -----------------------------------------
        # In the resulting df_grown, all df rows will be placed before all df_new rows,
        # because a duplicates fall into the range of df_new.
        # We intentionally keep df to remain in df_grown, 
        # because we want to keep datasets that were expensively generated from df
        # , although df is older than df_new.
        # It's demonstrated that if two rows are duplicates, the 1st row survives by drop_duplicates.
        # It's also demonstrated that if we concatenate two dataframes [df1, df2], df1's rows come first ...
        # WIERD, if we omit the column names here, there are no duplicate rows. Do they change data over time?
        df_grown = pd.concat([df, df_new]).drop_duplicates(subset=['Date', 'HomeTeam', 'AwayTeam'])
        assert len(df_grown) == len(df_new)

        # Find extra rows. We know all extra rows are placed after all df rows.
        df_extra = df_grown.iloc[df.shape[0]:]

        # Restore df ids for the df part and assign new ids to the df_extra part.
        ids = org_ids + list(range(max(list(org_ids))+1, max(list(org_ids))+1+df_extra.shape[0]))
        assert len(ids) == df_grown.shape[0]
        df_grown.insert(loc=0, column='id', value=ids)

        # df_extra can have any new ids. Preferably, assign them the same ids as in df_grown.
        ids = list(range(max(list(org_ids))+1, max(list(org_ids))+1+df_extra.shape[0]))
        assert len(ids) == df_extra.shape[0]
        df_extra.insert(loc=0, column='id', value=ids)

    except:
        print("find and grow raised an exception !!!")
        successful = False

    return successful, df_grown, df_extra

def assign_seasonal_filenames(countryFolder):
    xlsxFiles = []
    for (dirpath, dirnames, files) in os.walk(countryFolder):     # Make sure the .csv files are renamed .xlsx.
        for filename in files:
            if '.xlsx' in filename and '~' not in filename and '-' not in filename and 'df_' not in filename:
                xlsxFiles.append((dirpath, filename))
    print('files found to rename: ', len(xlsxFiles))

    total_rows = 0
    for (_dirPath, xlsxFilename) in xlsxFiles:
            filePath = os.path.join(_dirPath, xlsxFilename)
            # df = read_csv(filePath) doesn't work with football-data.co.uk data.
            df = read_excel(filePath)
            org = df.shape[0]
            df = improve_uk_dataframe(df, dropNa=False, dateOnly=True)
            newFilePath = os.path.join(_dirPath, os.path.splitext(xlsxFilename)[0].split(' ')[0] + '-' + str(min(df['Date']).year) + '-' + str(max(df['Date']).year)+'.xlsx')
            # print(newFilePath)
            os.rename(filePath, newFilePath)
            total_rows += df.shape[0]
            del df
    return


def improve_uk_dataframe(df, dropNa=True, dateOnly=False):
        # print("checking in improve_uk_dataframe: rows ", df.shape[0])
        if dropNa:
                #====================== Let this be the ONLY line where we lose some or many rows.============================
                df.dropna(subset=df.columns, inplace=True)
        else:   # fill in missing odds
                if dateOnly == False:

                        ogs = find_1X2_odds_group(df.columns)
                        og_cols = []
                        for og in ogs: og_cols += [og+'H', og+'D', og+'A']
                        
                        # norm_params = get_normalization_params(df, og_cols)     # We wish the cols has few nan values.
                        
                        bookies_with_missing_odds = []
                        for index, row in df.iterrows():
                                nn_og = None
                                for og in ogs:
                                        if not (pd.isnull(row[og+'H']) or pd.isnull(row[og+'D']) or pd.isnull(row[og+'A'])):
                                                # normal_h = (row[og+'H'] - norm_params[og+'H'][0]) / norm_params[og+'H'][1]
                                                # normal_d = (row[og+'D'] - norm_params[og+'D'][0]) / norm_params[og+'D'][1]
                                                # normal_a = (row[og+'A'] - norm_params[og+'A'][0]) / norm_params[og+'A'][1]
                                                nn_og = og; break
                                        
                                if nn_og is None: continue      # give up this raw. it will be removed soon

                                for og in ogs:  # replace keeping std and mean.
                                        if pd.isnull(row[og+'H']) or pd.isnull(row[og+'D']) or pd.isnull(row[og+'A']):
                                                bookies_with_missing_odds.append(og)
                                                df.loc[index, og+'H'] = row[nn_og+'H']; df.loc[index, og+'D'] = row[nn_og+'D']; df.loc[index, og+'A'] = row[nn_og+'A']
                                        # col = og+'H'
                                        # if pd.isnull(row[col]): (mean, std, max) = norm_params[col]; df.loc[index, col] = normal_h * std + mean
                                        # col = og+'D'
                                        # if pd.isnull(row[col]): (mean, std, max) = norm_params[col]; df.loc[index, col] = normal_d * std + mean
                                        # col = og+'A'
                                        # if pd.isnull(row[col]): (mean, std, max) = norm_params[col]; df.loc[index, col] = normal_a * std + mean
                        
                        bookies_with_missing_odds = list(set([(bookies_with_missing_odds.count(b), b) for b in bookies_with_missing_odds]))
                        bookies_with_missing_odds.sort(reverse=True)
                        print("Bookies with missing 1X2 odds: !!! filled ", bookies_with_missing_odds)

                        #====================== Let this be the ONLY line where we lose some or many rows.============================
                        df.dropna(subset=df.columns, inplace=True)

        df['Date'] = standardize_dates_uk(list(df['Date']))
        # Didn't convert to a uniform case, because football-data.uk team names have no case errors,
        # although a few teams names have a space at the end.
        df['HomeTeam'] = [team.strip() for team in [re.sub(r"\s", "_", item) for item in list(df['HomeTeam'])]]
        df['AwayTeam'] = [team.strip() for team in [re.sub(r"\s", "_", item) for item in list(df['AwayTeam'])]]
        # print("checking out improve_uk_dataframe: rows ", df.shape[0])

        return df

def standardize_date_uk(rugged_date):
        if isinstance(rugged_date, datetime.datetime):  # If Regional Format == English(United Kingdom), all calls fall to this branch.
                date = rugged_date
        elif isinstance(rugged_date, str):
                raise("Unexpected string date !!!!!! Check 'regional format.")
                
                try:
                        date = datetime.datetime.strptime(rugged_date, '%d/%m/%Y')
                except:
                        date = datetime.datetime.strptime(rugged_date, '%d/%m/%y')                                              
        else:
                print(rugged_date)
                raise Exception('unhandled date')
        
        if isinstance(date, datetime.datetime):
                date = date.date()

        return date

def standardize_dates_uk(rugged_dates):
        standardize_dates = [standardize_date_uk(d) for d in rugged_dates]
        return standardize_dates


def get_normalization_params(df, cols):
    def get_mean_and_std(col):
        array = df[col].dropna()
        array = np.array(array)
        return (array.mean(), array.std(), np.max(array))
    params = {}
    for col in cols:
        params[col] = get_mean_and_std(col)
    return params



def createGameGraph_uk(df, selectedDivs=None, selectedGames=None):
        graph = nx.Graph()

        for fixture_id, div, home, away, fixture_date in zip(list(df['id']),
                                                        list(df['Div']),
                                                        list(df['HomeTeam']), 
                                                        list(df['AwayTeam']),
                                                        # standardize_dates_uk(list(df['Date']))        # already standardized.
                                                        list(df['Date'])
                                                        ):
                
                if ((selectedDivs is not None) and (div not in selectedDivs)): continue
                if ((selectedGames is not None) and (fixture_id not in selectedGames)): continue

                if (home, away) not in graph.edges:
                        graph.add_edge(home, away)
                        # graph[home][away]['games'] = []
                        graph.edges[home, away]['games'] = []

                # graph[home][away]['games'].append((fixture_id, fixture_date))
                graph.edges[home, away]['games'].append((fixture_id, fixture_date, 0.0))        # 0.0 for conductance

        # This block removes duplicate fixture_ids in the data, like 563602.
        # This block reduces number of games significantly, from 474000 to 3xxxxx. Comment out
        for edge in graph.edges:
                # games = graph[edge[0]][edge[1]]['games']
                games = graph.edges[edge]['games']
                games = list(set(games))                # to remove redundance.
                # graph[edge[0]][edge[1]]['games'] = games
                graph.edges[edge]['games'] = games

        return graph
                
def fixture_id_to_ids_uk(
        df,
        gameGraph,
        a_conductance_search=0.1,
        b_conductance_search=1.5,
        conductance365=0.8, 
        conductanceMedTeam=0.2, 
        ids_max = 30,
        minConductanceStep = 0.002,
        renew=True,
        ):
               
        alpha = pow(conductance365, -1/365)
        minCon = 1e-9

        def createRegistanceView(gameGraph, baseId, baseHome, baseAway, baseDate, weight='resistance'):
                def conductance(_date):
                        return pow(alpha, (_date - baseDate).days)

                # calculate gameGraph[.][.]['resistance'] based on base_date
                # cnt_if = 0; cnt_else = 0
                for (teamA, teamB) in gameGraph.edges:  # generats unique edges: both (1, 2) and (2, 1) are represented as (1, 2) 
                        con = minCon
                        for (_id, _date) in gameGraph[teamA][teamB]['games']:
                                if(baseDate <= _date):
                                        # print('if', cnt_if, cnt_else, _id, baseId)
                                        pass
                                        # cnt_if += 1      
                                else:   # _id != baseId and _date <= baseDate
                                        # print('else', cnt_else, _id, baseId)
                                        if(teamA==baseHome and teamB==baseAway or teamB==baseHome and teamA==baseAway):
                                                # the game's two teams are baseHome and baseAway. The conductance is the standard.
                                                con += conductance(_date)     
                                        else:   
                                                # the game has a team that is neither baseHome nor baseAway. The conductance is less than the standard.
                                                con += 1 / (1 / conductanceMedTeam + 1 / conductance(_date) )
                                        # cnt_else += 1
                        gameGraph[teamA][teamB][weight] = 1/con
                        # print('con: ', con, 1/con, cnt_if, cnt_else)

                return gameGraph
        
        def getPossibleUniquePairsFromDijkstras(homeDijkstra, awayDijkstra):
                reachablesHome = list(set(list(homeDijkstra.keys())))
                pairsHome = [(teamA, teamB) for teamA in reachablesHome for teamB in reachablesHome if teamA < teamB]
                reachablesAway = list(set(list(awayDijkstra.keys())))
                pairsAway = [(teamA, teamB) for teamA in reachablesAway for teamB in reachablesAway if teamA < teamB]

                pairs = list(set(pairsHome + pairsAway))

                return pairs
      
        def getGamesFaster(gameGraph, possibleUniquePairs, baseId, baseDate):
                pairs = [(A, B) for (A, B) in possibleUniquePairs if (A, B) in gameGraph.edges()]
                dates_ids = []
                for (A, B) in pairs:
                        dates_ids += [(date, id) for (id, date) in gameGraph[A][B]['games'] if date < baseDate]
                # dates_ids = [((date, id) for (id, date) in gameGraph[A][B]['games'] if date <= baseDate and id != baseId) for (A, B) in pairs]
                # print(dates_ids)
                dates_ids.sort(reverse=True)
                ids_dates = [id for (ts, id) in dates_ids]
                # print(ids_dates)
                return ids_dates
        
        def getPair(df, baseId):
                frame = df.loc[df['id'] == baseId, ['HomeTeam', 'AwayTeam']]
                pairs = list(zip(list(frame['HomeTeam']), list(frame['AwayTeam'])))
                return pairs[0]
        
        def getGameSequence(gameGraph, baseEdge, baseId, baseDate, pathMinConductivity):
                # # I believe len(games), where game is the return of this function, is a monotonically decreasing function of pathMinConductivity.
                # gameGraph = createRegistanceView(gameGraph, baseId, baseEdge[0], baseEdge[1], baseDate, weight='resistance')
                homeDijkstra = nx.single_source_dijkstra_path(gameGraph, baseEdge[0], cutoff = 1/pathMinConductivity, weight='resistance')
                awayDijkstra = nx.single_source_dijkstra_path(gameGraph, baseEdge[1], cutoff = 1/pathMinConductivity, weight='resistance')
                # print('reachables: ', len(list(homeDijkstra.keys())), len(list(awayDijkstra.keys())))

                possibleUniquePairs = getPossibleUniquePairsFromDijkstras(homeDijkstra, awayDijkstra)
                games = getGamesFaster(gameGraph, possibleUniquePairs, baseId, baseDate)
                # games = getGames(df, possibleUniquePairs, baseId, baseDate)
                # print('ids: ', len(games))
                return games
        
        def sort_id_to_ids(id_to_ids):
                dates_ids = [(id_to_ids[id][0], id) for id in id_to_ids.keys()]
                dates_ids.sort()
                # print('dates_ids', dates_ids)
                id_to_ids = {str(id): id_to_ids[id][1] for (ts, id) in dates_ids}
                # print('ids: ', id_to_ids)
                return id_to_ids

                # frame = df.loc[df['fixture_id'] == baseId, ['teams_home_team_id', 'teams_away_team_id']]
                # pairs = list(zip(list(frame['teams_home_team_id']), list(frame['teams_away_team_id'])))
                # return pairs[0]

        count = 0
        fixture_id_to_ids = {} 
        for baseEdge in list(gameGraph.edges):  #-----------------------------------------------------------------------------
                baseGames = gameGraph[baseEdge[0]][baseEdge[1]]['games']
                # print('1. baseGames', baseGames)

                for (baseId, baseDate) in baseGames:        # required, as diffrent games have different dates leading to diffrernt conductivity
                        # if baseId != 576033: continue   #---------------------------------------

                        if fixture_id_to_ids.get(baseId, None) is not None:
                                pass
                                # print("baseEdge: {}, baseId: {}".format(baseEdge, baseId))
                                # raise("data error: duplicate texture_id found:")
                        else:
                                gameGraph = createRegistanceView(gameGraph, baseId, baseEdge[0], baseEdge[1], baseDate, weight='resistance')
                                # Although know max conductivity of a single game is 1.0, we start from b_conductance_search of more than 1.0
                                # bacause multiple games between a home and a away might increase conductance between them to more than 1.0.
                                a = a_conductance_search; b = b_conductance_search
                                games_a = getGameSequence(gameGraph, baseEdge, baseId, baseDate, a); ids_a = len(games_a)
                                # print(games_a)
                                games_b = getGameSequence(gameGraph, baseEdge, baseId, baseDate, b); ids_b = len(games_b)
                                # print(games_b)

                                while True:
                                        # print(ids_a, ids_b)
                                        if ids_a <= ids_max or ids_a == ids_b:  # no hope
                                                games = games_a
                                                break
                                        elif ids_b >= ids_max:  # no hope
                                                games = games_b
                                                break
                                        else: # find optimal pathMinConductivity in (a, b)
                                                c = (a + b) / 2
                                                # print('c = ', c)
                                                if c - a <= minConductanceStep:
                                                        # if ids_a < ids_b:
                                                        #         games = games_b
                                                        # else:
                                                        #         games = games_a

                                                        if abs(ids_max-ids_a) <= abs(ids_max-ids_b):
                                                                games = games_a
                                                        else:
                                                                games = games_b
                                                        break
                                                else:   # find in (a, b). we know ids_a > ids_max, idx_a != ids_b, ids_max > ids_b
                                                        games_c = getGameSequence(gameGraph, baseEdge, baseId, baseDate, c)
                                                        ids_c = len(games_c)
                                                        # print(ids_a, ids_b, ids_c)
                                                        if ids_a > ids_max and ids_max > ids_c: # we have a hope
                                                                b = c; ids_b = ids_c; games_b = games_c
                                                        elif ids_c >= ids_max and ids_max > ids_b:      # we have a hope
                                                                a = c; ids_a = ids_c; games_a = games_c
                                                        else:   # ids_max > idx_b
                                                                print('logical error')
                                                                # raise("logical error")
                                
                                                             
                                # games = getGameSequence(gameGraph, baseEdge, baseId, baseDate, pathMinConductivity)

                                if games is not None and len(games) >= 0:       # include no past-sequence games. 
                                        fixture_id_to_ids.update({baseId: (baseDate, games)})

                        count += 1
                        print("#.baseIds done: {}, #.keys done: {}, baseId: {}, baseDate: {}, pair: {}, len(games): {}                                  " \
                              .format(count, len(list(fixture_id_to_ids.keys())), baseId, baseDate, getPair(df, baseId), len(fixture_id_to_ids[baseId][1])), end='\r')

        fixture_id_to_ids = sort_id_to_ids(fixture_id_to_ids)

        return fixture_id_to_ids

#---------------------------------------------------------------------------------------------------------------------------------------------

def getGameSequence(gameGraph, baseEdge, baseId, baseDate, targetLength, minCurrent, conductance365, sinceDaysAgo, qualityPct):     # gameGraph IS changed.

        def create_Conductance_View(gameGraph, baseDate, conductance365, focusPairs=None):     
                alpha = pow(conductance365, -1/365)
                def conductance(_date):
                        daysAgo = (baseDate - _date).days       # > 0
                        return pow(alpha, -daysAgo)             # pow(conductance365, daysAgo/365) <= 1, as conductance365 < 1
                
                pairsToRemove = []
                def find_conductance(teamA, teamB):
                        edge_cond = 0.0
                        new_games = []
                        for game in gameGraph.edges[teamA, teamB]['games']:
                                _id = game[0]; _date = game[1]  # game[2] may or may not exist.
                                if(_date < baseDate):
                                        sub_cond = conductance(_date)
                                        edge_cond += sub_cond  # +, as they are parallel conductances
                                        new_games.append((_id, _date, sub_cond))
                                else:
                                        new_games.append((_id, _date, 0.0))     # Mask with zero conductivity !!!

                        if len(new_games) <= 0:
                                pairsToRemove.append((teamA, teamB))
                        else:
                                gameGraph.edges[teamA, teamB]['games'] = new_games
                                gameGraph.edges[teamA, teamB]['conductance'] = edge_cond

                        return

                if focusPairs is None:
                        for (teamA, teamB) in gameGraph.edges:  # generats unique edges: both (1,2) and (2,1) are represented as either (1,2) or (2,1) 
                                find_conductance(teamA, teamB)
                else:
                        for (teamA, teamB) in focusPairs:
                                find_conductance(teamA, teamB)

                for pair in pairsToRemove:      gameGraph.remove_edge(pair[0], pair[1])

                return gameGraph

        
        def find_Electric_Flow_On_Connected_Graph(graph, source, target):
                '''
                Find the flow of electric current on each edge of 'graph' when total flow of 1.0 flows from 'source' to 'target' nodes,
                with conductance of edges stored in edge['conductance'].
                1. If 'source' and 'target' nodes are disconnected with each other, WEIRD amounts on flows on edges.
                2. If 'source' and 'target' nodes have zero conductance between them, it produces WEIRD amount of flows that violate Kirchhoff's law.
                '''
                edges = [e for e in graph.edges]        # (u, v) either u < v or u < v.
                nodes = [v for v in graph.nodes]
                edges_signs = []
                for v in nodes:
                        # We want the orientation of an edge (u, v) to be from min(u, v) to max(u, v)
                        edges_v = [(v, u) for u in nodes if (v, u) in graph.edges]      # all edges that have v as its node, in the form of (v,u)
                        edges_v_plus = [(v, u) for (v, u) in edges_v if v > u]          # (v, u < v)
                        edges_v_minus = [(v, u) for (v, u) in edges_v if v < u]         # (v, u >= v)
                        edges_v_plus = [(edges.index((u, v)) if edges.count((u, v)) > 0 else edges.index((v, u))) for (u, v) in edges_v_plus]
                        edges_v_minus = [(edges.index((u, v)) if edges.count((u, v)) > 0 else edges.index((v, u))) for (u, v) in edges_v_minus]
                        edges_signs.append((edges_v_plus, edges_v_minus))
                matrix_B = [ [ (1 if e in edges_signs[v][1] else -1 if e in edges_signs[v][0] else 0) for e in range(len(edges)) ] for v in range(len(nodes))]
                matrix_B = np.array(matrix_B, dtype=np.float32)
                vector_C = np.array( [graph.edges[e]['conductance'] for e in edges], dtype=np.float32)
                matrix_C = np.diag(vector_C)
                matrix_L = np.dot(np.dot(matrix_B, matrix_C), matrix_B.T)
                inverse_L = np.linalg.pinv(matrix_L, hermitian=True)    # Moore-Penrose pseudo-inverse
                source_node = nodes.index(source)
                target_node = nodes.index(target)
                X_vector = np.zeros((len(nodes),), dtype=np.float32)
                X_vector[source_node] = 1.0
                X_vector[target_node] = -1.0
                flow = np.matmul(matrix_C, matrix_B.T)
                flow = np.matmul(flow, inverse_L)
                flow = np.matmul(flow, X_vector)

                return flow, nodes, edges   # nodes and edges, just in case list(graph.nodes) might have different order each time.
        
        def try_remove_lowest_flow_games(gameGraph, eFlow, edges, targetLength, qualityPct, minCurrent):
                # print('5 entering removal. nTotal: ', find_nTotalGames(gameGraph))
                emptyPairs = []
                currents = []
                nTotalGames = 0
                for (teamA, teamB) in gameGraph.edges:  # May not: teamA < teamB
                        eId = (edges.index((teamA, teamB)) if edges.count((teamA, teamB)) > 0 else edges.index((teamB, teamA))) # Exists
                        flow = abs(eFlow[eId])
                        if len(gameGraph.edges[teamA, teamB]['games']) <= 0:
                                emptyPairs.append((teamA, teamB))
                        else:
                                edge_conductance = gameGraph.edges[teamA, teamB]['conductance'] # Asserted positive, no need epsilon.
                                currentPerUnitCon = flow / edge_conductance
                                pair_games = gameGraph.edges[teamA, teamB]['games']     # Creates a set of pair expressions.
                                pair_current = [(currentPerUnitCon * cond, id, date, cond, teamA, teamB) for (id, date, cond) in pair_games]
                                currents += pair_current

                for (teamA, teamB) in emptyPairs:   gameGraph.remove_edge(teamA, teamB)

                currents.sort() # default: increasing order of the 1st axis, which is the flow amount.
                nTotalGames = len(currents)

                # We don't remove games if their total number is less than targetLengh, even though they may have low flows.
                if len(currents) > targetLength:
                        mark = 0; currLen = len(currents)
                        for idx in range(currLen):
                                if currents[idx][0] < minCurrent and nTotalGames - idx - 1 >= targetLength:     # mark can be idx + 1
                                        mark = idx + 1
                                else: break
                        currents = currents[mark:]

                # We don't remove games if their total number is less than targetLengh, even though they may have low flows.
                if len(currents) > targetLength:
                        #========= Remove low-flow games from 'currents'
                        nToRemove = max(1, round((len(currents)-targetLength) * (100-qualityPct)/100 ))
                        currents = currents[nToRemove:]         # remove low-current games
                        nGoals = len(currents)

                pairsChanged = []
                if find_nTotalGames(gameGraph) > len(currents):
                        nGoal = len(currents)

                        # print('9, initial len(currents), after nToRemove removal: ', len(currents))
                        #-----------------------------------------------------------------------------------------------------
                        #   Below, 'currents' is reflected to gameGraph. No more pairs/games are removed, except that.
                        #-----------------------------------------------------------------------------------------------------

                        #======== Find <which games on which pair> are in 'currents'
                        pairsToKeep = list(set([(teamA, teamB) for (_,_,_,_, teamA, teamB) in currents]))
                        gamesByPair = [ ( (teamA, teamB), [(id, date, con) for (_, id, date, con, _teamA, _teamB) in currents 
                                if _teamA == teamA and _teamB == teamB ] )      # currents and pairs_from_current share the same expressions of pair.
                                for (teamA, teamB) in pairsToKeep]       # May not: teamA < teamB

                        #========= Remove existing pairs that are not in pairsToKeep. 
                        allPairs = [(teamA, teamB) for (teamA, teamB) in gameGraph.edges]   # Creates another set of pair expressions.
                        # 'and' just in case the graph changes its pair expressoin.
                        pairsToRemove = [(teamA, teamB) for (teamA, teamB) in allPairs if (teamA, teamB) not in pairsToKeep and (teamB, teamA) not in pairsToKeep]
                        for (teamA, teamB) in pairsToRemove:  gameGraph.remove_edge(teamA, teamB)

                        #========= Replace existing games of pairsToKeep with games found in 'currents' if appropriate.  
                        for ((teamA, teamB), games) in gamesByPair:
                                if len(gameGraph.edges[teamA, teamB]['games']) != len(games):   # if some games were excluded.
                                        gameGraph.edges[teamA, teamB]['games'] = games  # Replace.
                                        pairsChanged.append((teamA, teamB))
                
                        #========= Update nTotalGames
                        nTotalGames = find_nTotalGames(gameGraph)
                        if nTotalGames != nGoal:
                                i = 3

                # print('10, returning from removal. nTotalGames: ', nTotalGames)

                return gameGraph, nTotalGames, pairsChanged, currents
        
        
        def collecAllGames(gameGraph):
                games = []
                for edge in gameGraph.edges:
                        games += [(date, id) for (id, date, cond) in gameGraph.edges[edge]['games']]

                games.sort(reverse=True)
                games = [int(id) for (_, id) in games]
                return games
        
        def collectGamesOnCurrent(ordered_current, targetLength):
                # ordered_current: (current * cond, id, date, cond, teamA, teamB) - increasingly ordered on current
                trimmed_ordered_current = ordered_current[-targetLength:]       # Save large currents.
                games = [(date, id) for (_, id, date, _, _, _) in trimmed_ordered_current]

                games.sort(reverse=True)
                games = [int(id) for (_, id) in games]
                return games
        
        def collectGamesOnConductance(conductanceView, targetLength):
                conductance = []
                nTotalGames = 0
                for (teamA, teamB) in conductanceView.edges:
                        # print('3', gameGraph.edges[teamA, teamB])
                        pair_games = conductanceView.edges[teamA, teamB]['games']
                        # print('10', 'eId', eId)
                        pair_conductance = [(cond, id, date) for (id, date, cond) in pair_games]
                        conductance += pair_conductance
                conductance.sort()      # increasing order of conductance
                trimmed_ordered_conductance = conductance[-targetLength:]       # Save large conductance
                games = [(date, id) for (_, id, date) in trimmed_ordered_conductance]

                games.sort(reverse=True)
                games = [int(id) for (_, id) in games]
                return games

        def find_nTotalGames(gameGraph):
                nTotalGames = 0
                for e in gameGraph.edges:
                        nTotalGames += len(gameGraph.edges[e]['games'])
                return nTotalGames
        
        def reachable(graph, u, v):
                _reachable = True
                try:
                        nx.shortest_path_length(graph, u, v)
                except:
                        _reachable = False
                return _reachable

        #--------------------------------------------------------------------------------------------------------------------

        # This block may speed up, but brings inconsitency when refining existing id_to_ids.
        # By commenting out this block, we allow pairs whit both nodes connected none of the base teams, at this stage of call.
        # PairsToKeep = get_All_Pairs_With_Both_Ends_Connected_To_At_Least_One_Base_Team(gameGraph, baseEdge, baseId, baseDate)
        # gameGraph = trim_GameGraph_By_Removing_Unnecessary_Pairs(gameGraph, PairsToKeep)  # deepcopy gameGraph

        # print('6', baseId, 'stat after edge trimming', getStat(gameGraph))
        # print('7', baseId, 'stat before future removing', getStat(gameGraph))
        gameGraph = remove_Long_Past_And_Future_Games(gameGraph, baseDate, sinceDaysAgo)
        # print('13', len(collecAllGames(gameGraph)))

        # print('8', baseId, 'stat after future removing', getStat(gameGraph))

        nTotalGames = find_nTotalGames(gameGraph)
        # print('seg. time layer: baseId, nTotalGames:', baseId, nTotalGames)
        games = None; changedPairs = None; ordered_current = None
        early_stopped = False
        tag = 0         #---------- tag 0: normal -------------
        # print('3', nTotalGames, targetLength)
        while nTotalGames > targetLength:
                # print('4', baseId, nTotalGames)
                gameGraph = create_Conductance_View(gameGraph, baseDate, conductance365, focusPairs=changedPairs)
                # print('seq. while1: after condview baseId, nTotalGames:', baseId, find_nTotalGames(gameGraph))
                if baseEdge[0] not in gameGraph.nodes or baseEdge[1] not in gameGraph.nodes:
                        early_stopped = True; tag = 1; break
                if not reachable(gameGraph, baseEdge[0], baseEdge[1]):
                # if baseEdge[0] not in [n for n in gameGraph.neighbors(baseEdge[1])]:
                        early_stopped = True; tag = 2; break      #---- because find_Electric_Flow_On_Connected_Graph gives weird/unreasonable eFlow for a disconnected graph.
                try:    # linalg: getting the inverse matrix may fail.
                        # Flow exists even if conductance is zero if graph is connected.
                        eFlow, nodes, edges = find_Electric_Flow_On_Connected_Graph(gameGraph, baseEdge[0], baseEdge[1])
                except:
                        early_stopped = True; tag = 3; break

                nTotalGames = find_nTotalGames(gameGraph)
                # print('seq. while2: before removal. baseId, nTotalGames:', baseId, nTotalGames)

                # It does NOT remove any if the # edges are equal or greater than targetLength.
                gameGraph, nTotalGames, changedPairs, ordered_current \
                = try_remove_lowest_flow_games(gameGraph, eFlow, edges, targetLength, qualityPct, minCurrent)
                nTotalGames = find_nTotalGames(gameGraph)
                # print('3', baseId, nTotalGames)

        # print('seq. while finished. tag: ', tag)

        label = 0
        if early_stopped:
                if ordered_current is not None:
                        # print('12')
                        games = collectGamesOnCurrent(ordered_current, targetLength); label = 1
                else:
                        # print('13')
                        games = collectGamesOnConductance(gameGraph, targetLength); label = 2
        else:
                # print('14')
                games = collecAllGames(gameGraph); label = 0

        # print('seq. leaving. label: ', label, len(games))

        return games, tag, label, gameGraph    # expected: list of (id, date)

def get_All_Pairs_With_Both_Ends_Connected_To_At_Least_One_Base_Team(gameGraph, baseEdge, baseId, baseDate):
        A_neibours = [node for node in gameGraph.neighbors(baseEdge[0])]
        B_neibours = [node for node in gameGraph.neighbors(baseEdge[1])]
        neibours = list(set(A_neibours + B_neibours))
        possibleUniquePairs = [(teamA, teamB) for teamA in neibours for teamB in neibours if teamA < teamB]
        pairs = [(A, B) for (A, B) in possibleUniquePairs if (A, B) in gameGraph.edges()]

        return pairs

def trim_GameGraph_By_Removing_Unnecessary_Pairs(gameGraph, PairsToKeep):
        edgesToRemove = [(u, v) for (u, v) in gameGraph.edges if ((u, v) not in PairsToKeep and (v, u) not in PairsToKeep)]
        for (u, v) in edgesToRemove:
                gameGraph.remove_edge(u, v)
        return gameGraph

def remove_Long_Past_And_Future_Games(gameGraph, baseDate, sinceDaysAgo):
        for (teamA, teamB) in gameGraph.edges:
                # print('13', teamA, teamB)
                games = gameGraph.edges[teamA, teamB]['games']
                # print('14', len(games))
                games = [(id, date, cond) for (id, date, cond) in games 
                        if ((baseDate-datetime.timedelta(days=sinceDaysAgo)) < date and date < baseDate)]
                # print('15', len(games))
                gameGraph.edges[teamA, teamB]['games'] = games
        return gameGraph

def sort_id_to_ids(id_to_ids):
        dates_ids = [(baseId, report, games) for (baseId, (report, games)) in id_to_ids.items()]
        dates_ids.sort()        # increasing on baseId
        # print('dates_ids', dates_ids)
        id_to_ids = {int(baseId): (report, games) for (baseId, report, games) in dates_ids}
        # print('ids: ', id_to_ids)
        return id_to_ids

        # frame = df.loc[df['fixture_id'] == baseId, ['teams_home_team_id', 'teams_away_team_id']]
        # pairs = list(zip(list(frame['teams_home_team_id']), list(frame['teams_away_team_id'])))
        # return pairs[0]

def getPair(df, baseId):
        frame = df.loc[df['id'] == baseId, ['HomeTeam', 'AwayTeam']]
        pairs = list(zip(list(frame['HomeTeam']), list(frame['AwayTeam'])))
        return pairs[0]

def save_step_id_to_ids(path, step_id_to_ids, work_id_to_ids, old_step):
        save = {}
        if old_step >= 0:
                if len(work_id_to_ids) > 0:   # Don't save anew unless we have extra id_to_ids, because this saving sometimes saves a wrong file.
                        save = step_id_to_ids | sort_id_to_ids(work_id_to_ids)
                        SaveJsonData(save, path)
                else:
                        save = step_id_to_ids
        return save

def fixture_id_to_ids_uk_maxflow(countryDirPath, id_to_ids_filename, targetLength, _minEdgeCond, sinceDaysAgo, _qualityPct, _conductance365, df_total, df_search, chooseDivs=False):                
        # df_total
        # df_search
        
        count = 0
        nFails = 0; divs = None
        nMissings = 0
        tag0 = tag1 = tag2 = tag3 = label0 = label1 = label2 = 0

        # df_search = df_total.iloc[-100:]     #----------------------------------------- test. Remove it.

        # get based on df_search
        id_list = list(df_search['id']); div_list = list(df_search['Div']); home_list = list(df_search['HomeTeam']); away_list = list(df_search['AwayTeam']); date_list = list(df_search['Date'])
        # print('22', home_list)
        search = zip(id_list, home_list, away_list, date_list)

        # get based on df_total
        _gameGraph = createGameGraph_uk(df_total, selectedDivs=None)   # None for all Divisions.

        step_size = int(1E3)        # do not change.
        old_step = -1
        total_id_to_ids = {}
        step_id_to_ids = {}
        work_id_to_ids = {}

        for baseId, home, away, baseDate in search:
                """
                # Consider adopting multiprocessing for each step, though this code doesn't work in ipynb.
                work = (["A", 5], ["B", 2], ["C", 1], ["D", 3])
                def work_log(work_data):
                print(" Process %s waiting %s seconds" % (work_data[0], work_data[1]))
                time.sleep(int(work_data[1]))
                print(" Process %s Finished." % work_data[0])
                def pool_handler():
                p = Pool(2)
                p.map(work_log, work)
                if __name__ == '__main__':
                pool_handler()
                """

                step = int(baseId/step_size) * step_size   # sure step >= 0

                def build_path(step):
                        return os.path.join(countryDirPath, '_id_to_ids', id_to_ids_filename + '-step-' + str(step) + '-size-' + str(step_size) + '.json')

                # Note the final step is always not saved. Save it after this loop.
                if step != old_step:    # We are turning to a new step.
                        save = save_step_id_to_ids(build_path(old_step), step_id_to_ids, work_id_to_ids, old_step)
                        total_id_to_ids = total_id_to_ids | save
                        # print('total ', len(total_id_to_ids))
                        step_id_to_ids = {}

                        path = build_path(step)
                        id_to_ids_read = LoadJsonData(path)
                        if id_to_ids_read is not None:
                                step_id_to_ids = id_to_ids_read
                        work_id_to_ids = {}

                        old_step = step
                
                count += 1
                if str(baseId) in step_id_to_ids.keys(): continue

                baseEdge = (home, away)

                if chooseDivs:
                        #=================== This was an attempt to speed up, by restricting search area to the divisions where the home and away teams attend.
                        #=================== But even if the very home and away teams don't attent directly, their linked teams may attend to divisions.
                        #=================== So, no restriction!
                        home_idx = list(set([i for i, x in enumerate(home_list) if x == home] + [i for i, x in enumerate(away_list) if x == home]))
                        away_idx = list(set([i for i, x in enumerate(home_list) if x == away] + [i for i, x in enumerate(away_list) if x == away]))
                        idx = list(set(home_idx + away_idx))
                        divs = list(set([div_list[id] for id in idx]))
                        # get based on df_total
                        gameGraph = createGameGraph_uk(df_total, selectedDivs=divs)
                else:
                        gameGraph = copy.deepcopy(_gameGraph)

                games, tag, label, _ = getGameSequence(gameGraph, baseEdge, baseId, baseDate, targetLength, _minEdgeCond, _conductance365, sinceDaysAgo, _qualityPct)
                
                if tag == 0: tag0 += 1
                elif tag == 1: tag1 += 1
                elif tag == 2: tag2 += 1
                elif tag == 3: tag3 += 1
                else: raise Exception("Unknown tag")

                if label == 0: label0 += 1
                elif label == 1: label1 += 1
                elif label == 2: label2 += 1
                else: raise Exception("Unknown label")

                if games is not None and len(games) >= 0:       # include no past-sequence games.
                        # print('111', baseId, baseDate, tag, games)
                        work_id_to_ids[baseId] = (baseDate, tag, label, games)
                        # print('112', fixture_id_to_ids[baseId])    
                        nMissings += max(0, (targetLength - len(games)))
                        print("baseId: {}, baseDate: {}, nGames: {}, tag: {}, label: {}, accTags: {}, accLabels: {}, nMissings: {}, selDivs: {}                          " \
                        .format(baseId, baseDate, len(games), tag, label, (tag0, tag1, tag2, tag3), (label0, label1, label2), nMissings, divs), end='\r')
                else:
                        nFails += 1

        # Give a chance to the final step to save.
        save = save_step_id_to_ids(build_path(old_step), step_id_to_ids, work_id_to_ids, old_step)
        total_id_to_ids = total_id_to_ids | save

        return total_id_to_ids

def CREATE_MAP_v1(folder, idMap_filename, targetLength, df_sequence, df_base, year_span, testcount=-1, to_save=True):

        def find_Electric_Flow_On_Connected_Graph(graph, source, target, inpuCurrent):
                '''
                Find the flow of electric current on each edge of 'graph' when total flow of 1.0 flows from 'source' to 'target' nodes,
                with conductance of edges stored in edge['conductance'].
                1. If 'source' and 'target' nodes are disconnected with each other, WEIRD amounts on flows on edges.
                2. If 'source' and 'target' nodes have zero conductance between them, it produces WEIRD amount of flows that violate Kirchhoff's law.
                '''
                edges = [e for e in graph.edges]        # (u, v) either u < v or u < v.
                nodes = [v for v in graph.nodes]
                edges_signs = []
                for v in nodes:
                        # We want the orientation of an edge (u, v) to be from min(u, v) to max(u, v)
                        edges_v = [(v, u) for u in nodes if (v, u) in graph.edges]      # all edges that have v as its node, in the form of (v,u)
                        edges_v_plus = [(v, u) for (v, u) in edges_v if v > u]          # (v, u < v)
                        edges_v_minus = [(v, u) for (v, u) in edges_v if v < u]         # (v, u >= v)
                        edges_v_plus = [(edges.index((u, v)) if edges.count((u, v)) > 0 else edges.index((v, u))) for (u, v) in edges_v_plus]
                        edges_v_minus = [(edges.index((u, v)) if edges.count((u, v)) > 0 else edges.index((v, u))) for (u, v) in edges_v_minus]
                        edges_signs.append((edges_v_plus, edges_v_minus))
                matrix_B = [ [ (1 if e in edges_signs[v][1] else -1 if e in edges_signs[v][0] else 0) for e in range(len(edges)) ] for v in range(len(nodes))]
                matrix_B = np.array(matrix_B, dtype=np.float32)
                vector_C = np.array( [graph.edges[e]['conductance'] for e in edges], dtype=np.float32)
                matrix_C = np.diag(vector_C)
                matrix_L = np.dot(np.dot(matrix_B, matrix_C), matrix_B.T)
                inverse_L = np.linalg.pinv(matrix_L, hermitian=True)    # Moore-Penrose pseudo-inverse
                source_node = nodes.index(source)
                target_node = nodes.index(target)
                X_vector = np.zeros((len(nodes),), dtype=np.float32)
                X_vector[source_node] = inpuCurrent
                X_vector[target_node] = - inpuCurrent
                flows = np.matmul(matrix_C, matrix_B.T)
                flows = np.matmul(flows, inverse_L)
                flows = np.matmul(flows, X_vector)

                return flows, nodes, edges   # nodes and edges, just in case list(graph.nodes) might have different order each time.

        def find_nTotalGames(gameGraph):
                nTotalGames = 0
                for e in gameGraph.edges:   nTotalGames += len(gameGraph.edges[e]['games'])
                return nTotalGames

        def reachable(graph, u, v):
                _reachable = True
                try:    nx.shortest_path_length(graph, u, v)
                except: _reachable = False
                return _reachable

        def isConnected(gameGraph):
                connected = True
                for u in gameGraph.nodes:
                        for v in gameGraph.nodes:
                                if not reachable(gameGraph, u, v):
                                        connected = False
                                        break
                return connected

        def create_conducting_game_graph_uk(game_list, baseDate, conductance365):
                # game_list is sorted in (date, div).
                graph = nx.Graph()
                alpha = pow(conductance365, -1/365)
                def conductance(_date):
                        daysAgo = (baseDate - _date).days       # > 0
                        return pow(alpha, -daysAgo)             # pow(conductance365, daysAgo/365) <= 1, as conductance365 < 1

                for id, div, home, away, dt in game_list:
                        if (home, away) not in graph.edges:
                                graph.add_edge(home, away)
                                graph.edges[home, away]['games'] = []
                        graph.edges[home, away]['games'].append((id, dt, conductance(dt)))

                for edge in graph.edges:
                        games = graph.edges[edge]['games']
                        edge_con = 0.0
                        for (_, _, con) in games: edge_con += con
                        graph.edges[edge]['conductance'] = edge_con

                return graph


        def find_games(gameGraph):
                games = []
                for e in gameGraph.edges:   games += [id for (id, date, con) in gameGraph.edges[e]['games']]
                return games

        def try_remove_lowest_flow_games(gameGraph, eFlows, edges, targetLength):
                # Eithe isConnected(gameGraph) or not. This call worsens it by removing some edges.
                emptyPairs = []
                currents = []
                for (teamA, teamB) in gameGraph.edges:  # May not: teamA < teamB
                        # get the current on the edge (teamA, teamB)
                        eId = (edges.index((teamA, teamB)) if edges.count((teamA, teamB)) > 0 else edges.index((teamB, teamA))) # Exists
                        flow = abs(eFlows[eId])
                        
                        if len(gameGraph.edges[teamA, teamB]['games']) <= 0:
                                emptyPairs.append((teamA, teamB))
                        else:
                                edge_conductance = gameGraph.edges[teamA, teamB]['conductance'] # Asserted positive, no need epsilon.
                                currentPerUnitCon = flow / edge_conductance
                                pair_games = gameGraph.edges[teamA, teamB]['games']     # Creates a set of pair expressions.
                                pair_current = [(currentPerUnitCon * cond, id, date, cond, teamA, teamB) for (id, date, cond) in pair_games]
                                currents += pair_current

                # Note: Pair representations (A, B) in currents came from [for (A, B) in gameGraph.edges]
                for (teamA, teamB) in emptyPairs:   gameGraph.remove_edge(teamA, teamB)
                assert find_nTotalGames(gameGraph) == len(currents)

                # sort currents in (curr / descending, date / descending )
                currents = [(date, curr, id, cond, teamA, teamB) for (curr, id, date, cond, teamA, teamB) in currents]
                currents.sort(reverse=True)     # later dates come first
                currents = [(curr, id, date, cond, teamA, teamB) for (date, curr, id, cond, teamA, teamB) in currents]
                currents.sort(reverse=True)    # larger current comes first

                if len(currents) <= targetLength:   pass
                else:
                        # Either zero-current edges survive or positive-current edges are removed by this cut, both leading to disconnected graph. 
                        currents = currents[ : targetLength ]   # note currents are sorted in (curr, date)
                        pairsChanged = []
                        #-----------------------------------------------------------------------------------------------------
                        #   Below, 'currents' is reflected to gameGraph. No more pairs/games are removed, except that.
                        #-----------------------------------------------------------------------------------------------------

                        #======== Find <which games on which pair> are in 'currents'
                        pairsInCurrents = list(set([(teamA, teamB) for (_,_,_,_, teamA, teamB) in currents]))
                        gamesByPairInCurrents = [ ( (teamA, teamB), [(id, date, con) for (_, id, date, con, _teamA, _teamB) in currents 
                                if _teamA == teamA and _teamB == teamB ] )      # currents and pairs_from_current share the same expressions of pair.
                                for (teamA, teamB) in pairsInCurrents]       # May not: teamA < teamB

                        #========= Remove existing pairs that are not in pairsInCurrents, that has no game at all in currents.
                        allPairs = [(teamA, teamB) for (teamA, teamB) in gameGraph.edges]   # Pair representatoin comes from [for (A, B) in gameGraph.edges]
                        pairsToRemove = [(teamA, teamB) for (teamA, teamB) in allPairs if (teamA, teamB) not in pairsInCurrents]
                        for (teamA, teamB) in pairsToRemove:  gameGraph.remove_edge(teamA, teamB)

                        #========= Replace existing games of pairsInCurrents with games found in 'currents' if appropriate.  
                        for ((teamA, teamB), games) in gamesByPairInCurrents:
                                if len(gameGraph.edges[teamA, teamB]['games']) != len(games):   # if some games were excluded.
                                        gameGraph.edges[teamA, teamB]['games'] = games  # Replace.
                                        pairsChanged.append((teamA, teamB))

                        #========= Update nTotalGames
                        nTotalGames = find_nTotalGames(gameGraph)
                        assert nTotalGames == targetLength

                return gameGraph, nTotalGames, pairsChanged, currents


        def get_historical_games_intra_div(base_id, base_date, base_div, home, away, div_sub_list, history_len, inputCurrent, conductance365):
                games = []; report = 0
                if len(div_sub_list) <= history_len:
                        games = [id for (id, _, _, _, _) in div_sub_list]
                        report = 10
                else:
                        dgg = create_conducting_game_graph_uk(div_sub_list, base_date, conductance365=conductance365)   # MUCH faster than transform_to_conducting_graph

                        if reachable(dgg, home, away):
                                flows, nodes, edges = find_Electric_Flow_On_Connected_Graph(dgg, home, away, inputCurrent)
                                flows_copy = copy.deepcopy(flows)
                                flows_copy = [abs(f) for f in flows_copy]
                                flows_copy.sort(reverse=True)

                                if  flows_copy[0] > inputCurrent/100:
                                        # Now, either isConnected(dgg) or not. This call worsens it by removing some edges.
                                        dgg, nTotalGames, pairsChanged, currents = try_remove_lowest_flow_games(dgg, flows, edges, history_len)
                                        games = find_games(dgg)
                                        assert len(games) == history_len
                                        report = 20
                                else:
                                        games = [id for (id, div, home, away, dt) in div_sub_list[- history_len : ]]  # collect latest ids
                                        report = 30

                        else: # Very rare. Few games are new edge in the conductance graph after collecting at lease HISTORY_LEN past games in the graph. They might be inter-league games.
                                # find_Electric_Flow_On_Connected_Graph(.) doesn't work here.
                                games = [id for (id, _,_,_,_) in div_sub_list[- history_len : ]]   # collect latest ids
                                report = 40

                return games, report
        
             
     
        def get_historical_games(base_id, base_date, base_div, home, away, sub_list, history_len, inputCurrent, conductance365):
                games = []; report = None
                if len(sub_list) <= history_len:
                        games = [id for (id, _, _, _, _) in sub_list]      # better than dummy games.
                        report = 100
                else:
                        gg = create_conducting_game_graph_uk(sub_list, base_date, conductance365=conductance365)   # MUCH faster than transform_to_conducting_graph

                        if reachable(gg, home, away):
                                flows, nodes, edges = find_Electric_Flow_On_Connected_Graph(gg, home, away, inputCurrent)
                                flows_abs = copy.deepcopy(flows)
                                flows_abs = [abs(f) for f in flows_abs]
                                flows_abs.sort(reverse=True)

                                if  flows_abs[0] > inputCurrent/100: # No other way to find the e flow problem was successful.
                                        # Now, either isConnected(dgg) or not. This call worsens it by removing some edges.
                                        gg, nTotalGames, pairsChanged, currents = try_remove_lowest_flow_games(gg, flows, edges, history_len)
                                        games = find_games(gg)
                                        assert len(games) == history_len
                                        report = 200
                                else:
                                        # either isConnected(gg) or not.
                                        div_sub_list = [(id, div, home, away, dt) for (id, div, home, away, dt) in sub_list if div == base_div]
                                        games, _report = get_historical_games_intra_div(base_id, base_date, base_div, home, away, div_sub_list, history_len, inputCurrent, conductance365)
                                        report = 300 + _report
                                        
                        else: # Very rare. Few games are new edge in the conductance graph after collecting at lease HISTORY_LEN past games in the graph. They might be inter-league games.
                                # find_Electric_Flow_On_Connected_Graph(.) doesn't work here.
                                div_sub_list = [(id, div, home, away, dt) for (id, div, home, away, dt) in sub_list if div == base_div]
                                games, _report = get_historical_games_intra_div(base_id, base_date, base_div, home, away, div_sub_list, history_len, inputCurrent, conductance365)
                                report = 400 + _report

                        if len(games) < history_len:
                                candi_games = [id for (id, _, _, _, _) in sub_list if id not in games]
                                games = games + candi_games[ : history_len - len(games)]        #--------------- ERROR, not latest games.
                                report += 1     # 200, 311, 320, 420, 330

                        assert len(games) == history_len

                        # if not isConnected(gg): report += 1000         # expensive

                games.sort(reverse=True)
                return games, report
        
        def sort_id_to_ids(id_to_ids):
                dates_ids = [(baseId, report, games) for (baseId, (report, games)) in id_to_ids.items()]
                dates_ids.sort()        # increasing on baseId
                # print('dates_ids', dates_ids)
                id_to_ids = {int(baseId): (report, games) for (baseId, report, games) in dates_ids}
                return id_to_ids
        
        def save_step_id_to_ids(path, step_id_to_ids, work_id_to_ids, old_step, to_save):
                save = {}
                if old_step >= 0:
                        if len(work_id_to_ids) > 0:   # Don't save anew unless we have extra id_to_ids, because this saving sometimes saves a wrong file.'"Electrical Flows 2.pdf"
                                save = step_id_to_ids | sort_id_to_ids(work_id_to_ids)
                                if to_save: SaveJsonData(save, path)
                        else:
                                save = step_id_to_ids
                return save

        #=========================================================================== Main =======================================================================

        id_list = list(df_sequence['id']); div_list = list(df_sequence['Div']); home_list = list(df_sequence['HomeTeam']); away_list = list(df_sequence['AwayTeam']); date_list = list(df_sequence['Date'])
        total_list = list(zip(id_list, div_list, home_list, away_list, date_list))

        id_list_s = list(df_base['id']); div_list_s = list(df_base['Div']); home_list_s = list(df_base['HomeTeam']); away_list_s = list(df_base['AwayTeam']); date_list_s = list(df_base['Date'])
        search_list = list(zip(id_list_s, div_list_s, home_list_s, away_list_s, date_list_s))

        step_size = int(1E3)        # do not change.
        old_step = -1
        total_id_to_ids = {}
        step_id_to_ids = {}
        work_id_to_ids = {}

        # df_built = df_built.sort_values(['Date', 'Div'], ascending=[True, True])
        
        max_days_covered = 0; count = 0
        for (base_id, base_div, home, away, base_date) in search_list:      # date: yyyy-mm-dd        , in search_list
                if count == testcount: break
                count += 1

                step = int(base_id/step_size) * step_size   # sure step >= 0

                def build_path(step):
                        return os.path.join(folder, idMap_filename + '-step-' + str(step) + '-size-' + str(step_size) + '.json')

                # Note the final step is always not saved. Save it after this loop.
                if step != old_step:    # We are turning to a new step.
                        save = save_step_id_to_ids(build_path(old_step), step_id_to_ids, work_id_to_ids, old_step, to_save)
                        total_id_to_ids = total_id_to_ids | save
                        step_id_to_ids = {}
                        path = build_path(step)
                        id_to_ids_read = LoadJsonData(path)
                        if id_to_ids_read is not None:
                                step_id_to_ids = id_to_ids_read
                        work_id_to_ids = {}
                        old_step = step
                
                if str(base_id) in step_id_to_ids.keys():  continue

                #------------------------------------------------------------------------------------------------------- Goal: get games.
                #????????????????????????????????????? Shall we limit the list to max 5 years ?????????????????????????????????????????????????
                day_span = year_span * 365
                sub_list = [(id, div, home, away, dt) for (id, div, home, away, dt) in total_list if id < base_id and (base_date-dt).days <= day_span]
                # div_sub_list = [(id, div, home, away, dt) for (id, div, home, away, dt) in sub_list if div == base_div]

                inputCurrent = 1000.0
                games, report = get_historical_games(base_id, base_date, base_div, home, away, sub_list, targetLength, inputCurrent, conductance365=0.9)

                if len(games) > 0: days_covered = (base_date - date_list[id_list.index(games[-1])]).days
                else: days_covered = 0
                if max_days_covered < days_covered: max_days_covered = days_covered

                print("base_id: {}, report: {}, days_span: {}, games[:10]: {}" \
                      .format(base_id, report, days_covered, games[:10]), end='\r')
                #-------------------------------------------------------------------------------------------------------

                if len(games) >= 0: work_id_to_ids[base_id] = (report, games)

        # Give a chance to the final step to save.
        save = save_step_id_to_ids(build_path(old_step), step_id_to_ids, work_id_to_ids, old_step, to_save)
        total_id_to_ids = total_id_to_ids | save

        # print(len(total_id_to_ids))
        total_id_to_ids = { id : value for (id, value) in total_id_to_ids.items() if int(id) in id_list_s }

        return total_id_to_ids


def CREATE_MAP_v2(folder, idMap_filename, targetLength, df_sequence, df_base, year_span, testcount=-1, to_save=True):

        def find_Electric_Flow_On_Connected_Graph(graph, source, target, inpuCurrent):
                '''
                Find the flow of electric current on each edge of 'graph' when total flow of 1.0 flows from 'source' to 'target' nodes,
                with conductance of edges stored in edge['conductance'].
                1. If 'source' and 'target' nodes are disconnected with each other, WEIRD amounts on flows on edges.
                2. If 'source' and 'target' nodes have zero conductance between them, it produces WEIRD amount of flows that violate Kirchhoff's law.
                '''
                edges = [e for e in graph.edges]        # (u, v) either u < v or u < v.
                nodes = [v for v in graph.nodes]
                edges_signs = []
                for v in nodes:
                        # We want the orientation of an edge (u, v) to be from min(u, v) to max(u, v)
                        edges_v = [(v, u) for u in nodes if (v, u) in graph.edges]      # all edges that have v as its node, in the form of (v,u)
                        edges_v_plus = [(v, u) for (v, u) in edges_v if v > u]          # (v, u < v)
                        edges_v_minus = [(v, u) for (v, u) in edges_v if v < u]         # (v, u >= v)
                        edges_v_plus = [(edges.index((u, v)) if edges.count((u, v)) > 0 else edges.index((v, u))) for (u, v) in edges_v_plus]
                        edges_v_minus = [(edges.index((u, v)) if edges.count((u, v)) > 0 else edges.index((v, u))) for (u, v) in edges_v_minus]
                        edges_signs.append((edges_v_plus, edges_v_minus))
                matrix_B = [ [ (1 if e in edges_signs[v][1] else -1 if e in edges_signs[v][0] else 0) for e in range(len(edges)) ] for v in range(len(nodes))]
                matrix_B = np.array(matrix_B, dtype=np.float32)
                vector_C = np.array( [graph.edges[e]['conductance'] for e in edges], dtype=np.float32)
                matrix_C = np.diag(vector_C)
                matrix_L = np.dot(np.dot(matrix_B, matrix_C), matrix_B.T)
                inverse_L = np.linalg.pinv(matrix_L, hermitian=True)    # Moore-Penrose pseudo-inverse
                source_node = nodes.index(source)
                target_node = nodes.index(target)
                X_vector = np.zeros((len(nodes),), dtype=np.float32)
                X_vector[source_node] = inpuCurrent
                X_vector[target_node] = - inpuCurrent
                flows = np.matmul(matrix_C, matrix_B.T)
                flows = np.matmul(flows, inverse_L)
                flows = np.matmul(flows, X_vector)

                return flows, nodes, edges   # nodes and edges, just in case list(graph.nodes) might have different order each time.

        def find_nTotalGames(gameGraph):
                nTotalGames = 0
                for e in gameGraph.edges:   nTotalGames += len(gameGraph.edges[e]['games'])
                return nTotalGames

        def reachable(graph, u, v):
                _reachable = True
                try:    nx.shortest_path_length(graph, u, v)
                except: _reachable = False
                return _reachable

        def isConnected(gameGraph):
                connected = True
                for u in gameGraph.nodes:
                        for v in gameGraph.nodes:
                                if not reachable(gameGraph, u, v):
                                        connected = False
                                        break
                return connected

        def create_conducting_game_graph_uk(game_list, baseDate, conductance365):
                # game_list is sorted in (date, div).
                graph = nx.Graph()
                alpha = pow(conductance365, -1/365)
                def conductance(_date):
                        daysAgo = (baseDate - _date).days       # > 0
                        return pow(alpha, -daysAgo)             # pow(conductance365, daysAgo/365) <= 1, as conductance365 < 1

                for id, div, home, away, dt in game_list:
                        if (home, away) not in graph.edges:
                                graph.add_edge(home, away)
                                graph.edges[home, away]['games'] = []
                        graph.edges[home, away]['games'].append((id, dt, conductance(dt)))

                for edge in graph.edges:
                        games = graph.edges[edge]['games']
                        edge_con = 0.0
                        for (_, _, con) in games: edge_con += con
                        graph.edges[edge]['conductance'] = edge_con

                return graph


        def find_games(gameGraph):
                games = []
                for e in gameGraph.edges:   games += [id for (id, date, con) in gameGraph.edges[e]['games']]
                return games

        def try_remove_lowest_flow_games(gameGraph, eFlows, edges, targetLength):
                """
                1. Eithe isConnected(gameGraph) or not. This call worsens connectivity by removing some edges.
                2. Very small e-current incidently created by numerical errors unexpectedly affect this function.
                """ 
                emptyPairs = []
                currents = []
                for (teamA, teamB) in gameGraph.edges:  # May not: teamA < teamB. PairExpression_1 is created here.
                        # get the current on the edge (teamA, teamB)
                        eId = (edges.index((teamA, teamB)) if edges.count((teamA, teamB)) > 0 else edges.index((teamB, teamA))) # Exists.
                        flow = abs(eFlows[eId]) # May be very small. NP.
                        
                        if len(gameGraph.edges[teamA, teamB]['games']) <= 0:
                                emptyPairs.append((teamA, teamB))       # PairExpression_1
                        else:   # Enlist all games no matter how small e-current flows on them.
                                edge_conductance = gameGraph.edges[teamA, teamB]['conductance'] # Asserted positive, no need epsilon.
                                currentPerUnitCon = flow / edge_conductance
                                pair_games = gameGraph.edges[teamA, teamB]['games']     # Creates a set of pair expressions. Should be PairExpression_1, if deterministic.
                                pair_current = [(currentPerUnitCon * cond, id, date, cond, teamA, teamB) for (id, date, cond) in pair_games] # PairExpression_1
                                currents += pair_current

                # Note: Pair representations (A, B) in currents came from [for (A, B) in gameGraph.edges]
                for (teamA, teamB) in emptyPairs:   gameGraph.remove_edge(teamA, teamB)
                assert find_nTotalGames(gameGraph) == len(currents)

                # sort currents in (curr / descending, date / descending )
                currents = [(date, curr, id, cond, teamA, teamB) for (curr, id, date, cond, teamA, teamB) in currents]
                currents.sort(reverse=True)     # later dates come first
                currents = [(curr, id, date, cond, teamA, teamB) for (date, curr, id, cond, teamA, teamB) in currents]
                currents.sort(reverse=True)    # larger current comes first. PairExpression_1

                if len(currents) <= targetLength:   pass        # This doens't happen becasue we enlisted ALL games above.
                else:
                        # Either zero-current edges survive or positive-current edges are removed by this cut, both leading to disconnected graph. 
                        currents = currents[ : targetLength ]   # note currents are sorted in (curr, date)
                        pairsChanged = []
                        #-----------------------------------------------------------------------------------------------------
                        #   Below, 'currents' is reflected to gameGraph. No more pairs/games are removed, except that.
                        #-----------------------------------------------------------------------------------------------------

                        #======== Find <which games on which pair> are in 'currents'
                        pairsInCurrents = list(set([(teamA, teamB) for (_,_,_,_, teamA, teamB) in currents]))   # PairExpression_1
                        gamesByPairInCurrents = [ ( (teamA, teamB), [(id, date, con) for (_, id, date, con, _teamA, _teamB) in currents # PairExpression_1
                                if _teamA == teamA and _teamB == teamB ] )      # currents and pairs_from_current share the same expressions of pair.
                                for (teamA, teamB) in pairsInCurrents]       # May not: teamA < teamB

                        #========= Remove existing pairs that are not in pairsInCurrents, that has no game at all in currents.
                        allPairs = [(teamA, teamB) for (teamA, teamB) in gameGraph.edges]   # PairExpression_2 is created here.
                        pairsToRemove = [(teamA, teamB) for (teamA, teamB) in allPairs if ((teamA, teamB) not in pairsInCurrents and (teamB, teamA) not in pairsInCurrents)]
                        for (teamA, teamB) in pairsToRemove:  gameGraph.remove_edge(teamA, teamB)

                        #========= Replace existing games of pairsInCurrents with games found in 'currents' if appropriate.
                        pairsChanged = []
                        for ((teamA, teamB), games) in gamesByPairInCurrents: # PairExpression_1
                                if len(gameGraph.edges[teamA, teamB]['games']) != len(games):   # if some games were excluded.
                                        gameGraph.edges[teamA, teamB]['games'] = games  # Replace.
                                        pairsChanged.append((teamA, teamB))

                        #========= Update nTotalGames
                        nTotalGames = find_nTotalGames(gameGraph)
                        assert nTotalGames == targetLength      # because we enlisted all games.

                return gameGraph, nTotalGames, pairsChanged, currents
        
        #-------------------------------------------------------------------------------------------------------------
        def get_teams_by_div(df_sequence):
                div_list = list(df_sequence['Div']); home_list = list(df_sequence['HomeTeam']); away_list = list(df_sequence['AwayTeam'])
                unique_divs = set(div_list)
                teams_by_div = {div: set([home_list[i] for i in range(len(home_list)) if div_list[i] == div]).union(set([away_list[i] for i in range(len(away_list)) if div_list[i] == div])) for div in unique_divs}

                teams_by_div = { div : set([team for team in home_list if div_list[home_list.index(team)]==div]).union(set([team for team in away_list if div_list[away_list.index(team)]==div])) for div in unique_divs }
                for div1 in unique_divs:
                        for div2 in unique_divs:
                                if div1 != div2:
                                        assert teams_by_div[div1].intersection(teams_by_div[div2]) == set()
                teams_by_div = { div: list(teams) for (div, teams) in teams_by_div.items()}
                return teams_by_div
        teams_by_div = get_teams_by_div(df_sequence)
        #--------------------------------------------------------------------------------------------------------------


        def get_historical_games_intra_div(base_id, base_date, base_div, home, away, div_sub_list, history_len, inputCurrent, conductance365):
                """
                Goal: Choose as many as, most relevant, games from div_sub_list
                """
                games = []; seq_type = 0
                if len(div_sub_list) <= history_len:
                        games = [id for (id, _, _, _, _) in div_sub_list]
                        seq_type = 10
                else:
                        dgg = create_conducting_game_graph_uk(div_sub_list, base_date, conductance365=conductance365)   # MUCH faster than transform_to_conducting_graph

                        if reachable(dgg, home, away):
                                flows, nodes, edges = find_Electric_Flow_On_Connected_Graph(dgg, home, away, inputCurrent)
                                flows_copy = copy.deepcopy(flows)
                                flows_copy = [abs(f) for f in flows_copy]
                                flows_copy.sort(reverse=True)

                                if  flows_copy[0] > inputCurrent/10000:   # e-current tach seems successful.
                                        # Now, either isConnected(dgg) or not. This call worsens it by removing some edges.
                                        dgg, _, _, _ = try_remove_lowest_flow_games(dgg, flows, edges, history_len)
                                        games = find_games(dgg)
                                        assert len(games) == history_len
                                        seq_type = 20
                                else:   # e-current tech failed.
                                        games = [id for (id, _, _, _, _) in div_sub_list[- history_len : ]]  # collect latest ids
                                        seq_type = 30

                        else: # Very rare. Few games are new edge in the conductance graph after collecting at lease HISTORY_LEN past games in the graph. They might be inter-league games.
                                # find_Electric_Flow_On_Connected_Graph(.) doesn't work here.
                                games = [id for (id, _,_,_,_) in div_sub_list[- history_len : ]]   # collect latest ids
                                seq_type = 40

                return games, seq_type
        
             
     
        def get_historical_games_v2(base_id, base_date, base_div, home, away, sub_list, history_len, inputCurrent, conductance365):
                games = []; report = None
                if len(sub_list) <= history_len:
                        games = [id for (id, _, _, _, _) in sub_list]      # better than dummy games.
                        seq_type = 1
                else:
                        # This logic is based on the emperical proof that a team plays in only a division. Divisions, as a set of teams, have no intersection.

                        div_sub_list = [(id, div, home, away, dt) for (id, div, home, away, dt) in sub_list if div == base_div]
                        games, seq_type = get_historical_games_intra_div(base_id, base_date, base_div, home, away, div_sub_list, history_len, inputCurrent, conductance365)

                        if len(games) < history_len:
                                candi_games = [id for (id, _, _, _, _) in sub_list if id not in games]  # We know sub_list is ascendigly sorted in date.
                                games = games + candi_games[ - (history_len - len(games)) : ]         # So this chooses latest games.
                                seq_type += 1
                        assert len(games) == history_len
                        # if not isConnected(gg): report += 1000         # expensive

                games.sort(reverse=True)        # latest games come first.
                return games, seq_type
        
        def sort_id_to_ids(id_to_ids):
                dates_ids = [(baseId, report, games) for (baseId, (report, games)) in id_to_ids.items()]
                dates_ids.sort()        # increasing on baseId
                # print('dates_ids', dates_ids)
                id_to_ids = {int(baseId): (report, games) for (baseId, report, games) in dates_ids}
                return id_to_ids
        
        def save_step_id_to_ids(path, step_id_to_ids, work_id_to_ids, old_step, to_save):
                save = {}
                if old_step >= 0:
                        if len(work_id_to_ids) > 0:   # Don't save anew unless we have extra id_to_ids, because this saving sometimes saves a wrong file.'"Electrical Flows 2.pdf"
                                save = step_id_to_ids | sort_id_to_ids(work_id_to_ids)
                                if to_save: SaveJsonData(save, path)
                        else:
                                save = step_id_to_ids
                return save

        #=========================================================================== Main =======================================================================

        id_list = list(df_sequence['id']); div_list = list(df_sequence['Div']); home_list = list(df_sequence['HomeTeam']); away_list = list(df_sequence['AwayTeam']); date_list = list(df_sequence['Date'])
        total_list = list(zip(id_list, div_list, home_list, away_list, date_list))

        id_list_s = list(df_base['id']); div_list_s = list(df_base['Div']); home_list_s = list(df_base['HomeTeam']); away_list_s = list(df_base['AwayTeam']); date_list_s = list(df_base['Date'])
        search_list = list(zip(id_list_s, div_list_s, home_list_s, away_list_s, date_list_s))

        step_size = int(1E3)        # do not change.
        old_step = -1
        total_id_to_ids = {}
        step_id_to_ids = {}
        work_id_to_ids = {}

        # df_built = df_built.sort_values(['Date', 'Div'], ascending=[True, True])
        
        max_days_covered = 0; count = 0
        for (base_id, base_div, home, away, base_date) in search_list:      # date: yyyy-mm-dd        , in search_list
                if count == testcount: break
                count += 1

                step = int(base_id/step_size) * step_size   # sure step >= 0

                def build_path(step):
                        return os.path.join(folder, idMap_filename + '-step-' + str(step) + '-size-' + str(step_size) + '.json')

                # Note the final step is always not saved. Save it after this loop.
                if step != old_step:    # We are turning to a new step.
                        save = save_step_id_to_ids(build_path(old_step), step_id_to_ids, work_id_to_ids, old_step, to_save)
                        total_id_to_ids = total_id_to_ids | save
                        step_id_to_ids = {}
                        path = build_path(step)
                        id_to_ids_read = LoadJsonData(path)
                        if id_to_ids_read is not None:
                                step_id_to_ids = id_to_ids_read
                        work_id_to_ids = {}
                        old_step = step
                
                if str(base_id) in step_id_to_ids.keys():  continue

                #------------------------------------------------------------------------------------------------------- Goal: get games.
                #????????????????????????????????????? Shall we limit the list to max 5 years ?????????????????????????????????????????????????
                day_span = year_span * 365
                sub_list = [(id, div, home, away, dt) for (id, div, home, away, dt) in total_list if id < base_id and (base_date-dt).days <= day_span]
                # div_sub_list = [(id, div, home, away, dt) for (id, div, home, away, dt) in sub_list if div == base_div]

                inputCurrent = 1000.0
                games, seq_type = get_historical_games_v2(base_id, base_date, base_div, home, away, sub_list, targetLength, inputCurrent, conductance365=0.9)

                if len(games) > 0: days_covered = (base_date - date_list[id_list.index(games[-1])]).days
                else: days_covered = 0
                if max_days_covered < days_covered: max_days_covered = days_covered

                print("base_id: {}, seq_type: {}, days_span: {}, games[:10]: {}" \
                      .format(base_id, seq_type, days_covered, games[:10]), end='\r')
                #-------------------------------------------------------------------------------------------------------

                if len(games) >= 0: work_id_to_ids[base_id] = (seq_type, games)

        # Give a chance to the final step to save.
        save = save_step_id_to_ids(build_path(old_step), step_id_to_ids, work_id_to_ids, old_step, to_save)
        total_id_to_ids = total_id_to_ids | save

        # print(len(total_id_to_ids))
        total_id_to_ids = { id : value for (id, value) in total_id_to_ids.items() if int(id) in id_list_s }

        return total_id_to_ids



def remove_folder_contents(folderPath):
    for (root, dirs, files) in os.walk(folderPath):
        if root == folderPath:
            for file in files:
                filePath = os.path.join(folderPath, file)
                os.remove(filePath)
            for dir in dirs:
                dirPath = os.path.join(folderPath, dir)
                shutil.rmtree(dirPath)

def copy_id_map(sourcePath, destPath, destNote):
    files = []
    for (root, dirs, files) in os.walk(sourcePath):     # Make sure the .csv files are renamed .xlsx.
        for file in files:
            splits = file.split('-'); splits[2] = destNote
            newFileName = "-".join(splits)
            print(newFileName)
            filePath = os.path.join(sourcePath, file)
            newfilePath = os.path.join(destPath, newFileName)
            shutil.copy(filePath, newfilePath)
    return

def copy_dataset(sourcePath, destPath, destNote):
    files = []
    # print(sourcePath, destPath, destNote)
    for (root, dirs, files) in os.walk(sourcePath):
        if root == sourcePath:
            for file in files:
                splits = file.split('-'); splits[2] = destNote
                newFileName = "-".join(splits)
                filePath = os.path.join(sourcePath, file)
                newfilePath = os.path.join(destPath, newFileName)
                shutil.copy(filePath, newfilePath)
            for dir in dirs:
                splits = dir.split('-'); splits[2] = destNote
                newFileName = "-".join(splits)
                filePath = os.path.join(sourcePath, dir)
                newfilePath = os.path.join(destPath, newFileName)
                shutil.copytree(filePath, newfilePath)

from tokenizers import (decoders, models, normalizers, pre_tokenizers, processors, trainers, Tokenizer)

def createSimpleTokenizer(corpus_files, vocab_size, unknown_token, special_tokens):
        tokenizer = Tokenizer(models.WordLevel(unk_token=unknown_token))
        tokenizer.normalizer = normalizers.Sequence([normalizers.NFD(), normalizers.Lowercase(), normalizers.StripAccents()])
        tokenizer.pre_tokenizer = pre_tokenizers.WhitespaceSplit()
        trainer = trainers.WordLevelTrainer(vocab_size=vocab_size, special_tokens=special_tokens)
        tokenizer.train(corpus_files, trainer=trainer)
        tokenizer.decoder = decoders.WordPiece(prefix=" ")
        return tokenizer

def creat_team_tokenizer_uk(master_df, tokenizer_folder_path):
    teams = list(set(list(master_df['HomeTeam']) + list(master_df['AwayTeam'])))
    teams_string = [str(team) for team in teams]
    teams_string = [re.sub(r"\s", "_", item) for item in teams_string]    # replace spaces with a '_'
    teams_text = " ".join(teams_string)

    corpus_file = os.path.join(tokenizer_folder_path, 'team_ids_text_uk.txt')
    f = open(corpus_file, "w+", encoding="utf-8")
    f.write(teams_text)
    f.close()

    corpus_files = [corpus_file]
    unknown_token = config['unknown_token']
    special_tokens = [unknown_token] ################### + ["[HOME]", "[AWAY]"]
    vocab_size = len(teams_string) + len(special_tokens)

    tokenizer_team = createSimpleTokenizer(corpus_files, vocab_size, unknown_token, special_tokens)
    return tokenizer_team

def convert_to_token(tokenizer, teamList):
    success = True
    teams = [team.strip() for team in [re.sub(r"\s", "_", item) for item in teamList]]
    teams = " ".join(teams)
    teams = tokenizer.encode(teams).tokens
    for id in range(len(teamList)):
        if teams[id] == config['unknown_token']:
            print("'", teamList[id], "' is unknown, and replaced with ", config['unknown_token'])
            success = False
    return teams, success

import difflib
def find_most_similar_teams(team_vocab, input, count):
    mr = difflib.SequenceMatcher()
    similarities = []
    for team in team_vocab:
        mr.set_seqs(team, input) 
        similarities.append((mr.ratio(), team))
    similarities.sort(reverse=True)
    similarities = similarities[:count]
    similarities = [team for (_, team) in similarities]
    return similarities

def convert_to_token_with_candidates(tokenizer, teamList, candi_count=3):
    team_vocab = list(tokenizer.get_vocab().keys())
    success = True
    teams = [team.strip() for team in [re.sub(r"\s", "_", item) for item in teamList]]
    teams = " ".join(teams)
    teams = tokenizer.encode(teams).tokens

    candidates = {}
    for id in range(len(teamList)):
        if teams[id] == config['unknown_token']:
            candidates = candidates | { teamList[id] : find_most_similar_teams(team_vocab, teamList[id], candi_count) }
            success = False
    return teams, candidates, success


def ensure_excelfile_exists(filePath) :
    try :
        if os.path.exists( filePath ) is True :
            workbook = openpyxl.load_workbook( filePath )
        else :
            workbook = openpyxl.Workbook()
            workbook.save( filePath )
    except :
        raise Exception( "Failed ensureing Excel file exists: {}".format(filePath) )
    return

def find_latest_bookie_dictionary(countryThemeFolder, divs):
    dictPath = os.path.join(countryThemeFolder, "df_train" + '.json')
    total_dict = LoadJsonData(dictPath)

    dicts_by_div = { div : [-1, {}] for div in divs }

    for (fileName, bookieDict) in total_dict.items():
            [div, yearA, _] = fileName.split('-')
            if dicts_by_div[div][0] < int(yearA):
                  dicts_by_div[div][0] = int(yearA)
                  dicts_by_div[div][1] = bookieDict

    dicts_by_div = { div : dict for (div, (maxYearA, dict)) in dicts_by_div.items() }

    first_dict = None
    for (div, dict) in dicts_by_div.items():
        if first_dict is None:
                first_dict = dict
        else:
            assert first_dict == dict

    return first_dict


def format_inference_input_excel_sheet(folder_path, filename, sheet, bookie_list):
    odds_cols = []
    for bookie in bookie_list:
        odds_cols += [bookie+'H', bookie+'D', bookie+'A']

    column_list = ['id', 'year', 'month', 'day', 'hour', 'minute', 'Div', 'HomeTeam', 'AwayTeam']
    input_columns = {col:[] for col in column_list}
    input_columns = input_columns | {col: pd.Series(dtype='float') for col in odds_cols}
    df = pd.DataFrame(input_columns)

    filePath = os.path.join(folder_path, filename)
    ensure_excelfile_exists(filePath)

    excelfile = pd.ExcelFile(filePath)
    sheetnames = excelfile.sheet_names
    intSheetNames = [int(name) for name in sheetnames if name.isdecimal()]
    digital_sheet = str(0) if len(intSheetNames) <= 0 else str(max(intSheetNames) + (1 if sheet.lower() == 'next' else 0))
    sheetName = (digital_sheet if (sheet.lower() == 'next' or sheet.lower() == 'same') else sheet)

    with pd.ExcelWriter(filePath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name=sheetName)
    return column_list, sheetName


def check_for_consistency_input_excel_sheet(countryTheme_folder_path, file_folder_path, filename, sheet, input_columns, tokenizer, bookie_list, divs):
    teams_by_div = get_teams_by_div(countryTheme_folder_path)

    filePath = os.path.join(file_folder_path, filename)
    df = pd.read_excel(filePath, sheet_name=sheet)
    if 'Unnamed: 0' in list(df.columns): df.drop('Unnamed: 0', axis=1, inplace=True)

    check_report = ""
    consistent = True
    # Check for consistency
    delta = list(set(input_columns) - set(df.columns))
    if len(delta) > 0: 
        consistent = False
        check_report += "Some columns are missing: {}\n".format(",".join(delta))
    if not (df.shape[0] > 0): 
        consistent = False
        check_report += "There are no data rows. \n"
    if df.isnull().any().any():
        consistent = False
        check_report += "There are some NaN values. \n"

    if set(['year', 'month', 'day', 'hour', 'minute']) <= set(df.columns):
        try:
            datetime_series = pd.to_datetime(df[['year', 'month', 'day', 'hour', 'minute']], utc=True)
            now = datetime.datetime.now(pytz.utc)
            delta_hours = [(dt - now).total_seconds()/3600 for dt in datetime_series]
            negatives = [id for id in range(len(delta_hours)) if delta_hours[id] < 0]
            if len(negatives) > 0:
                consistent = False
                check_report += "Some datetimes are past: rows = {}\n".format(",".join([str(id) for id in negatives]))
            aheads = [id for id in range(len(delta_hours)) if delta_hours[id] > 24]
            if max(delta_hours) > 24:
                consistent = False
                check_report += "Some datetimes are not within 24 hours: rows = {}\n".format(",".join([str(id) for id in aheads]))
        except:
            consistent = False
            check_report += "Some datetime columns have invalid data types or value range. \n"

    divs = list(df['Div']); homes = list(df['HomeTeam']); aways = list(df['AwayTeam'])
    for id in range(len(divs)):
           if homes[id] not in teams_by_div[divs[id]]:
                  consistent = False
                  check_report += "{} doesn't belong to {} teams\n".format(homes[id], divs[id])
           if aways[id] not in teams_by_div[divs[id]]:
                  consistent = False
                  check_report += "{} doesn't belong to {} teams\n".format(aways[id], divs[id])

    lists = [homes, aways]
    teamList = [val for tup in zip(*lists) for val in tup]
    teams, candidates, success = convert_to_token_with_candidates(tokenizer, teamList, candi_count=3)
    if success is False:
        consistent = False
        check_report += "Mis-spelled teams and their possible candidates come below: \n"
        for (spell, candis) in candidates.items():
            check_report += "\t{} : {}\n".format(spell, candis)

    delta = list(set(df['Div']) - set(divs))
    if len(delta) > 0:
        consistent = False
        check_report += "Some Div values are wrong: {}\n".format(delta)

    odds_cols = []
    for bookie in bookie_list:
        odds_cols += [bookie+'H', bookie+'D', bookie+'A']

    for col in odds_cols:
        odds = np.array(df[col])
        condition = odds <= 1.0
        if condition.any():
            consistent = False
            check_report += "Some odds are too small: rows {} columns {}\n".format(list(np.where(condition)[0]), col)
        condition = odds > 60.0
        if (condition).any():
            consistent = False
            check_report += "Some odds are too large: rows {} columns {}\n".format(list(np.where(condition)[0]), col)

    for bookie in bookie_list:
        cols = [bookie+'H', bookie+'D', bookie+'A']
        odds = np.array(df[cols])
        condition = np.sum(1/odds, axis=-1) > 1.2
        if condition:
            consistent = False
            check_report += "Some odds group are too small: odds group {}\n".format(cols)
        condition = np.sum(1/odds, axis=-1) < 0.85
        if condition:
            consistent = False
            check_report += "Some odds group are too large: odds group {}\n".format(cols)

    return consistent, check_report


def build_bbab_dataframe(file_folder_path, filename, sheet, bb_cols, ab_cols, bookie_dict):
    filePath = os.path.join(file_folder_path, filename)
    df = pd.read_excel(filePath, sheet_name=sheet)

    datetime_series = pd.to_datetime(df[['year', 'month', 'day', 'hour', 'minute']], utc=True)
    df.drop(['year', 'month', 'day', 'hour', 'minute'], axis=1, inplace=True)
    df.insert(3, 'Date', datetime_series.dt.date)
    df['id'] = range(config['baseGameId'] * 2, config['baseGameId'] * 2 + df.shape[0])
    
    dummy = [0 for _ in range(df.shape[0])]
    loc = 6
    for col in ab_cols:
        df.insert(loc, col, dummy)
        loc += 1

    rename_plan = {}
    for (k, v) in bookie_dict.items():
           rename_plan = rename_plan | {k+'H': v+'H', k+'D': v+'D', k+'A': v+'A'}
    df.rename(columns=rename_plan, inplace=True)

    df = df[bb_cols + ab_cols]

    return df

def get_subfolders(folder):
    subfolders = []
    for (root, dirs, files) in os.walk(folder):
        if root == folder:
            subfolders = dirs
            break
    return subfolders

def remove_subfolders(folder, subfolders):
    for sb in subfolders:
        path = os.path.join(folder, sb)
        shutil.rmtree(path)

def create_output_excel_sheet(history_folder, filename, sheetName, prediction, df_bbab_to_predict, bookie_dict, AB_cols, NUMBER_QUERIES):
        df_bbab_to_predict.drop(AB_cols, axis=1, inplace=True)

        rename_plan = {}
        for (k, v) in bookie_dict.items():
                rename_plan = rename_plan | {v+'H': k+'H', v+'D': k+'D', v+'A': k+'A'}
        df_bbab_to_predict.rename(columns=rename_plan, inplace=True)
        # print(df_bbab_to_predict)

        def oneHot_to_1X2(oneHot): return '1' if oneHot == [1,0,0] else ('X' if oneHot == [0,1,0] else '2')
        bookies = list(bookie_dict.keys())
        for (baseId, oneHots) in prediction.items():
                symbols = []
                for id in range(len(bookies)):
                        start = NUMBER_QUERIES * id
                        oneHot = oneHots[start: start+NUMBER_QUERIES]
                        symbols.append(oneHot_to_1X2(oneHot))
                prediction[baseId] = symbols
        # print(prediction)

        id_list = list(df_bbab_to_predict['id'])

        symbols_by_bookie = {}
        for id in range(len(bookies)):
                symbols_by_bookie = symbols_by_bookie | { bookies[id]: [ prediction[baseId][id] for baseId in id_list] }
        # print(symbols_by_bookie)

        for b in bookies:
                loc = list(df_bbab_to_predict.columns).index(b+'A')
                df_bbab_to_predict.insert(loc+1, b, symbols_by_bookie[b])
        # print(df_bbab_to_predict)

        cols = []
        for b in bookies:
                cols += [b+'H', b+'D', b+'A']
        df_bbab_to_predict.drop(cols, axis=1, inplace=True)

        filePath = os.path.join(history_folder, filename)
        sheetName = '_' + sheetName

        with pd.ExcelWriter(filePath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df_bbab_to_predict.to_excel(writer, sheet_name=sheetName)

        return

def find_start_and_end_dates(ds):
    baseDateStart = baseDateEnd = None
    for z in ds:
        _, _, _, _, _, baseDateDetails, _ = z
        if baseDateStart is None:
            detail = list(baseDateDetails.numpy()[0]) 
            baseDateStart = datetime.datetime(config['baseDate'].year + detail[0], detail[1], detail[2])
    detail = list(baseDateDetails.numpy()[0])
    baseDateEnd = datetime.datetime(config['baseDate'].year + detail[0], detail[1], detail[2])
    return baseDateStart, baseDateEnd

def linear_strategy(profits):
    balance_history = []; balance = 1.0
    max_balance = balance; deepest_canyon = max_balance - balance
    for i in range(len(profits)):
        balance += profits[i]   # Linear
        balance_history.append(balance)
        if max_balance < balance: max_balance = balance
        if deepest_canyon < max_balance - balance: deepest_canyon = max_balance - balance
    return balance_history, max_balance, deepest_canyon

def exponential_strategy(profits, safety):
    balance_history = []; balance = 1.0
    max_balance = balance; deepest_canyon = max_balance - balance
    for i in range(len(profits)):
        balance += profits[i] * max_balance / safety    # Exponential with safety
        balance_history.append(balance)
        if max_balance < balance: max_balance = balance
        if deepest_canyon < max_balance - balance: deepest_canyon = max_balance - balance
    return balance_history, max_balance, deepest_canyon