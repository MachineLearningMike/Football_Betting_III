
import numpy as np
import pandas as pd
# Wonder why both the two of the following lines are required.
import openpyxl
from openpyxl import load_workbook

import os
import shutil

class Excel_Driver() :

    def __init__( self ) :
        pass

    def Is_Writable_Type( self, value ) :

        return isinstance( value, int ) \
            or isinstance( value, float ) \
            or isinstance( value, np.float32 ) \
            or isinstance( value, str )


    def Backup( self, filePath, backup ) :

        try :
            # Make sure file exists.
            if os.path.exists( filePath ) is True :
                os.remove( backup ) if os.path.exists( backup ) else None
                shutil.copy( filePath, backup )
        except :
            raise( "Failture backing up {}".format( filePath ) )

        return


    def Ensure_File_Sheet_Exists( self, filePath, sheetName = None ) :

        try :
            # Make sure file exists.
            if os.path.exists( filePath ) is True :
                workbook = openpyxl.load_workbook( filePath )
            else :
                workbook = openpyxl.Workbook()
                workbook.save( filePath )
                workbook = openpyxl.load_workbook( filePath )

            # Make sure a sheet exists for this session.
            if sheetName is not None and sheetName not in workbook.sheetnames :
                workbook.create_sheet( title = sheetName )

            workbook.save( filePath )

        except :
            raise Exception( "Failed ensureing Excel sheet: {}.{}".format( filePath, sheetName if sheetName is not None else '' ) )

        return

    #======================== Write ========================

    def Save_Data_Dict( self, filePath, sheetName, dictionary, overwrite = False, horizontal = True, list_mode = False ) :

        sheet = None
        if os.path.exists( filePath ) is True :
            workbook = openpyxl.load_workbook( filePath )
            if sheetName in workbook.sheetnames :
                sheet = workbook[sheetName] # workbook.get_sheet_by_name( sheetName )

        if sheet is None :
            raise Exception( "File or sheet not found.".format( filePath, sheetName ) )

        try :
            if horizontal is True :
                self.Write_Dict_Horizontal( sheet, dictionary, overwrite )
            else :
                self.Write_Dict_Vertical( sheet, dictionary, overwrite, list_mode = list_mode )

            workbook.template = False
            workbook.save( filePath )

        except :
            raise Exception( "Failure writing to {}.{}".format( filePath, sheetName ) )       
    
        return


    def Save_Data_Dictionaries( self, filePath, sheetName, dictionaries, overwrite = False, horizontal = True, list_mode = False ) :

        sheet = None
        if os.path.exists( filePath ) is True :
            workbook = openpyxl.load_workbook( filePath )
            if sheetName in workbook.sheetnames :
                sheet = workbook[sheetName] # workbook.get_sheet_by_name( sheetName )

        if sheet is None :
            raise Exception( "File or sheet not found.".format( filePath, sheetName ) )

        try :
            for dictionary in dictionaries :
                if horizontal is True :
                    self.Write_Dict_Horizontal( sheet, dictionary, overwrite )
                else :
                    self.Write_Dict_Vertical( sheet, dictionary, overwrite, list_mode = list_mode )

            workbook.template = False
            workbook.save( filePath )

        except :
            raise Exception( "Failure writing to {}.{}".format( filePath, sheetName ) )       
    
        return

    def Write_Dict_Horizontal( self, sheet, dict, overwrite = False ) :

        cells_head_row = sheet[1]
        if overwrite : target_row = max( sheet.max_row, 2 )
        else : target_row = sheet.max_row + 1

        count = 0
        for key, value in dict.items() :

            head_cell_found = None
            for cell in cells_head_row :
                if cell.value is not None and cell.value == key :
                    head_cell_found = cell; break
            if head_cell_found is not None :
                if self.Is_Writable_Type( value ) :
                    sheet.cell( target_row, head_cell_found.column ).value = value
                else :
                    debug = 3
            else :
                max_column = sheet.max_column
                sheet.cell( 1, max_column + 1 ).value = key
                sheet.cell( target_row, max_column + 1 ).value = value

            count += 1

        print( "============= Total {} columns written to Excel file.".format( count ) )

        return


    def Write_Dict_Vertical( self, sheet, dict, overwrite = False, list_mode = False ) :

        def Write_Row( row_id, list_mode, value, target_col ):
            if list_mode :
                assert isinstance( value, list )
                for val_id in range( len( value ) ):
                    if self.Is_Writable_Type( value[val_id] ) :
                        sheet.cell( row_id, val_id + 2 ).value = value[val_id]
            else :
                assert not isinstance( value, list )
                if self.Is_Writable_Type( value ) :
                    sheet.cell( row_id, target_col ).value = value

        cells_head_column = sheet[ 'A' ]
        if overwrite : target_col = max( sheet.max_column, 2 )
        else : target_col = sheet.max_column + 1

        count = 0
        for key, value in dict.items() :

            head_cell_found = None
            for cell in cells_head_column :
                if cell.value is not None and cell.value == key :
                    head_cell_found = cell; break
            if head_cell_found is not None :
                Write_Row( head_cell_found.row, list_mode, value, target_col )
            else :
                max_row = sheet.max_row
                sheet.cell( max_row + 1, 1 ).value = key # column = 1 couples with 'A' above.
                Write_Row( max_row + 1, list_mode, value, target_col )

            count += 1

        print( "============= Total {} rows written to Excel file.".format( count ) )

        return

    #======================== Read ========================

    def Load_Data_Dict( self, filePath, sheetName, horizontal = True, reduce = 'last', list_mode = False ) :

        workbook = None
        if os.path.exists( filePath ) is True :
            workbook = openpyxl.load_workbook( filePath )
        
        if workbook is None :
            raise Exception( "File not found.".format( filePath ) )

        dictionary = None

        sheet = None
        if sheetName in workbook.sheetnames :
            sheet = workbook[sheetName] # workbook.get_sheet_by_name( sheetName )
        
        if sheet is not None :
            try :
                if horizontal is True :
                    dictionary = self.Read_Dict_Horizontal( sheet = sheet, reduce = reduce )
                else :
                    dictionary = self.Read_Dict_Vertical( sheet = sheet, reduce = reduce, list_mode = list_mode )
            except :
                raise Exception( "Failure reading from sheet {}".format( sheetName ) )
        else :
            dictionary = {}

        workbook.close()

        return dictionary


    def Read_Dict_Horizontal( self, sheet, reduce = 'last' ) :

        cells_head_row = sheet[1]
        target_row = sheet.max_row

        dictionary = {}

        count = 0
        for cell in cells_head_row :
            if cell.value is not None :
                value = sheet.cell( target_row, cell.column ).value
                if value is not None :
                    dictionary[ str( cell.value ) ] = value
                    count += 1

        print( "============= Total {} columns written to Excel file.".format( count ) )

        return


    def Read_Dict_Vertical( self, sheet, reduce = 'last', list_mode = False ) :

        cells_head_column = sheet[ 'A' ]
        target_col = sheet.max_column

        dictionary = {}

        if list_mode : 
            start_col = 2; end_col = sheet.max_column
        else :
            start_col = sheet.max_column; end_col = sheet.max_column

        count = 0
        for cell in cells_head_column :
            if cell.value is not None :
                val_list = []
                for col in range( start_col, end_col + 1 ) :
                    value = sheet.cell( cell.row, col ).value
                    if value is not None : val_list.append( value )

                if list_mode and len( val_list ) > 0 :
                    dictionary[ str( cell.value ) ] = val_list
                    count += 1 
                elif list_mode is False and len( val_list ) > 0 :
                    dictionary[ str( cell.value ) ] = val_list[0]
                    count += 1 

        print( "============= Total {} rows read from Excel file.".format( count ) )

        return dictionary


    def Write_DataFrames( self, frames_dict, filePath, sheetName, nFrames = -1 ) :

        writer = pd.ExcelWriter( filePath, engine = 'xlsxwriter' )
        workbook = writer.book
        worksheet = workbook.add_worksheet( sheetName )
        writer.sheets[ sheetName ] = worksheet

        prev_start_row = 1; prev_frame = None; countWritten = 0

        for frame_name, frame  in frames_dict.items() :

            row = prev_start_row + ( 1 if prev_frame is None else  prev_frame.shape[0] + 3 )
            
            worksheet.write_string( row, 0, frame_name )

            start_row = row + 2
            frame.to_excel( writer, sheet_name = sheetName, startrow = start_row, startcol = 0 )

            prev_start_row = start_row; prev_frame = frame

            countWritten += 1
            if 0 < nFrames and nFrames < countWritten : break

        writer.save()

        return